"""Resumable overnight account-summary processor.

Each Excel file is loaded through app_account.py and its existing account
summary download route. The resulting three sheets are stored in SQLite.
"""

from __future__ import annotations

import argparse
import gc
import json
import logging
import os
import sqlite3
import sys
import time
from contextlib import contextmanager, redirect_stderr, redirect_stdout
from io import BytesIO, StringIO
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

import app_account
from summary_database import (
    DEFAULT_DATABASE_PATH,
    connect_database,
    initialize_database,
    mark_file_failed,
    save_file_summaries,
    set_worker_state,
    utc_now,
)


LOGGER = logging.getLogger("account_summary_batch")
SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def configure_logging(database_path: Path, verbose: bool = False) -> None:
    database_path.parent.mkdir(parents=True, exist_ok=True)
    log_path = database_path.parent / "account_summary_worker.log"
    LOGGER.setLevel(logging.DEBUG if verbose else logging.INFO)
    LOGGER.handlers.clear()
    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    console = logging.StreamHandler()
    console.setFormatter(formatter)
    console.setLevel(logging.DEBUG if verbose else logging.INFO)
    LOGGER.addHandler(console)
    file_handler = RotatingFileHandler(
        log_path,
        maxBytes=10 * 1024 * 1024,
        backupCount=3,
        encoding="utf-8",
    )
    file_handler.setFormatter(formatter)
    file_handler.setLevel(logging.DEBUG)
    LOGGER.addHandler(file_handler)


class SingleWorkerLock:
    """Hold a non-blocking process lock next to the SQLite database."""

    def __init__(self, database_path: Path) -> None:
        self.path = database_path.with_suffix(database_path.suffix + ".worker.lock")
        self.handle: Any = None

    def __enter__(self) -> "SingleWorkerLock":
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.handle = self.path.open("a+b")
        self.handle.seek(0, os.SEEK_END)
        if self.handle.tell() == 0:
            self.handle.write(b"0")
            self.handle.flush()
        self.handle.seek(0)
        try:
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(self.handle.fileno(), msvcrt.LK_NBLCK, 1)
            else:
                import fcntl

                fcntl.flock(self.handle.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        except OSError as exc:
            self.handle.close()
            raise RuntimeError(
                "Another account-summary worker is already using this database."
            ) from exc
        return self

    def __exit__(self, exc_type: Any, exc: Any, traceback: Any) -> None:
        if not self.handle:
            return
        try:
            self.handle.seek(0)
            if os.name == "nt":
                import msvcrt

                msvcrt.locking(self.handle.fileno(), msvcrt.LK_UNLCK, 1)
            else:
                import fcntl

                fcntl.flock(self.handle.fileno(), fcntl.LOCK_UN)
        finally:
            self.handle.close()


@contextmanager
def keep_windows_awake(enabled: bool) -> Iterator[None]:
    if not enabled or os.name != "nt":
        yield
        return
    import ctypes

    es_continuous = 0x80000000
    es_system_required = 0x00000001
    ctypes.windll.kernel32.SetThreadExecutionState(
        es_continuous | es_system_required
    )
    try:
        yield
    finally:
        ctypes.windll.kernel32.SetThreadExecutionState(es_continuous)


def iter_excel_files(input_directory: Path) -> Iterator[Path]:
    for root, directory_names, file_names in os.walk(input_directory):
        directory_names[:] = sorted(
            name
            for name in directory_names
            if not name.startswith(".")
        )
        for file_name in sorted(file_names):
            if file_name.startswith("~$"):
                continue
            path = Path(root) / file_name
            if path.suffix.lower() in SUPPORTED_EXTENSIONS:
                yield path.resolve()


def discover_files(database_path: Path, input_directory: Path) -> dict[str, int]:
    """Register new/changed workbooks without disturbing completed results."""
    connection = connect_database(database_path)
    now = utc_now()
    counts = {"seen": 0, "new": 0, "changed": 0, "unchanged": 0}
    try:
        existing = {
            row["source_path"].casefold(): (row["id"], row["fingerprint"])
            for row in connection.execute(
                "SELECT id, source_path, fingerprint FROM source_files"
            )
        }
        with connection:
            for path in iter_excel_files(input_directory):
                try:
                    stat = path.stat()
                except OSError as exc:
                    LOGGER.warning("Could not inspect %s: %s", path, exc)
                    continue
                counts["seen"] += 1
                fingerprint = f"{stat.st_size}:{stat.st_mtime_ns}"
                existing_row = existing.get(str(path).casefold())
                if existing_row is None:
                    connection.execute(
                        """
                        INSERT INTO source_files (
                            source_path,
                            file_name,
                            file_size,
                            mtime_ns,
                            fingerprint,
                            status,
                            discovered_at,
                            last_seen_at
                        ) VALUES (?, ?, ?, ?, ?, 'pending', ?, ?)
                        """,
                        (
                            str(path),
                            path.name,
                            stat.st_size,
                            stat.st_mtime_ns,
                            fingerprint,
                            now,
                            now,
                        ),
                    )
                    counts["new"] += 1
                elif existing_row[1] != fingerprint:
                    connection.execute(
                        """
                        UPDATE source_files
                        SET file_name = ?,
                            file_size = ?,
                            mtime_ns = ?,
                            fingerprint = ?,
                            status = 'pending',
                            attempts = 0,
                            last_seen_at = ?,
                            error_message = NULL
                        WHERE id = ?
                        """,
                        (
                            path.name,
                            stat.st_size,
                            stat.st_mtime_ns,
                            fingerprint,
                            now,
                            existing_row[0],
                        ),
                    )
                    counts["changed"] += 1
                else:
                    counts["unchanged"] += 1
        return counts
    finally:
        connection.close()


def reset_interrupted_files(database_path: Path) -> int:
    connection = connect_database(database_path)
    try:
        with connection:
            cursor = connection.execute(
                """
                UPDATE source_files
                SET status = 'pending',
                    error_message = COALESCE(
                        error_message,
                        'Previous worker stopped during processing; queued again.'
                    )
                WHERE status = 'processing'
                """
            )
        return cursor.rowcount
    finally:
        connection.close()


def reset_failed_files(database_path: Path) -> int:
    connection = connect_database(database_path)
    try:
        with connection:
            cursor = connection.execute(
                """
                UPDATE source_files
                SET status = 'pending', attempts = 0, error_message = NULL
                WHERE status = 'failed'
                """
            )
        return cursor.rowcount
    finally:
        connection.close()


def reset_all_files(database_path: Path) -> int:
    connection = connect_database(database_path)
    try:
        with connection:
            cursor = connection.execute(
                """
                UPDATE source_files
                SET status = 'pending', attempts = 0, error_message = NULL
                """
            )
        return cursor.rowcount
    finally:
        connection.close()


def claim_next_file(
    database_path: Path,
    *,
    maximum_attempts: int,
) -> sqlite3.Row | None:
    connection = connect_database(database_path)
    try:
        connection.execute("BEGIN IMMEDIATE")
        row = connection.execute(
            """
            SELECT id, source_path, file_name, attempts
            FROM source_files
            WHERE (
                    status = 'pending'
                    OR (status = 'failed' AND attempts < ?)
                )
            ORDER BY
                CASE status WHEN 'pending' THEN 0 ELSE 1 END,
                id
            LIMIT 1
            """,
            (maximum_attempts,),
        ).fetchone()
        if row is None:
            connection.commit()
            return None
        connection.execute(
            """
            UPDATE source_files
            SET status = 'processing',
                attempts = attempts + 1,
                started_at = ?,
                completed_at = NULL,
                duration_seconds = NULL,
                error_message = NULL
            WHERE id = ?
            """,
            (utc_now(), row["id"]),
        )
        connection.commit()
        return row
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


def parse_summary_workbook(payload: bytes) -> dict[str, list[dict[str, Any]]]:
    expected_sheets = (
        "Account Wise Summary",
        "Bank Wise Summary",
        "Partial Bank Wise Summary",
    )
    workbook = load_workbook(
        BytesIO(payload),
        read_only=True,
        data_only=True,
    )
    try:
        summaries: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in expected_sheets:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(
                    f"Generated summary is missing sheet: {sheet_name}"
                )
            worksheet = workbook[sheet_name]
            row_iterator = worksheet.iter_rows(values_only=True)
            headers = next(row_iterator, None)
            if not headers:
                summaries[sheet_name] = []
                continue
            normalized_headers = [str(value).strip() for value in headers]
            summaries[sheet_name] = [
                dict(zip(normalized_headers, values))
                for values in row_iterator
                if any(value is not None for value in values)
            ]
        return summaries
    finally:
        workbook.close()


def analyse_with_existing_app(source_path: Path) -> dict[str, list[dict[str, Any]]]:
    """Call app_account.py's existing processing and report-generation code."""
    captured_output = StringIO()
    with redirect_stdout(captured_output), redirect_stderr(captured_output):
        app_account.uploaded_files_count = 0
        success, message = app_account.process_excel_file(
            str(source_path),
            is_first_file=True,
        )
        if not success:
            raise ValueError(message)
        with app_account.app.test_request_context(
            "/download-account-summary"
        ):
            response = app_account.download_account_summary()
            if isinstance(response, tuple):
                response = response[0]
            mimetype = getattr(response, "mimetype", "")
            response.direct_passthrough = False
            payload = response.get_data()
            if mimetype == "application/json":
                try:
                    error_data = json.loads(payload.decode("utf-8"))
                except (UnicodeDecodeError, json.JSONDecodeError):
                    error_data = {"message": payload[:500].decode(errors="replace")}
                raise ValueError(
                    error_data.get("message")
                    or error_data.get("error")
                    or "Account summary generation failed."
                )
    LOGGER.debug("app_account.py output for %s:\n%s", source_path, captured_output.getvalue())
    return parse_summary_workbook(payload)


def release_app_memory() -> None:
    app_account.df_main = None
    app_account.df_other_sheets = {}
    app_account.uploaded_files_count = 0
    app_account.debited_acc_map = {}
    app_account.credited_acc_map = {}
    app_account.debited_trans_id_map = {}
    app_account.credited_trans_id_map = {}
    app_account.breakdown_map = {}
    gc.collect()


def queue_counts(database_path: Path, maximum_attempts: int) -> dict[str, int]:
    connection = connect_database(database_path)
    try:
        row = connection.execute(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END)
                    AS completed,
                SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END)
                    AS failed,
                SUM(CASE
                    WHEN status = 'pending'
                      OR (status = 'failed' AND attempts < ?)
                    THEN 1 ELSE 0 END) AS actionable
            FROM source_files
            """,
            (maximum_attempts,),
        ).fetchone()
        return {key: int(row[key] or 0) for key in row.keys()}
    finally:
        connection.close()


def process_queue(
    database_path: Path,
    *,
    maximum_attempts: int,
    maximum_files: int | None,
) -> int:
    processed_this_run = 0
    while maximum_files is None or processed_this_run < maximum_files:
        source = claim_next_file(
            database_path,
            maximum_attempts=maximum_attempts,
        )
        if source is None:
            break
        source_path = Path(source["source_path"])
        started = time.perf_counter()
        set_worker_state(
            database_path,
            is_running=True,
            process_id=os.getpid(),
            current_file=source["file_name"],
            message="Analysing workbook",
        )
        try:
            if not source_path.is_file():
                raise FileNotFoundError(f"Source file no longer exists: {source_path}")
            summaries = analyse_with_existing_app(source_path)
            duration = time.perf_counter() - started
            counts = save_file_summaries(
                database_path,
                source["id"],
                summaries,
                duration_seconds=duration,
            )
            processed_this_run += 1
            progress = queue_counts(database_path, maximum_attempts)
            LOGGER.info(
                "[%s/%s] %s | %.2fs | %s ACK | %s account rows",
                progress["completed"] + progress["failed"],
                progress["total"],
                source["file_name"],
                duration,
                counts["acknowledgements"],
                counts["account"],
            )
        except KeyboardInterrupt:
            duration = time.perf_counter() - started
            mark_file_failed(
                database_path,
                source["id"],
                "Worker stopped by user during this file; it can be retried.",
                duration_seconds=duration,
            )
            raise
        except Exception as exc:
            duration = time.perf_counter() - started
            mark_file_failed(
                database_path,
                source["id"],
                f"{type(exc).__name__}: {exc}",
                duration_seconds=duration,
            )
            processed_this_run += 1
            LOGGER.exception(
                "Failed %s after %.2fs; continuing with the next file.",
                source_path,
                duration,
            )
        finally:
            release_app_memory()
    return processed_this_run


def build_argument_parser() -> argparse.ArgumentParser:
    default_input = Path(r"C:\Users\admin\Desktop\bank_trails")
    if not default_input.exists():
        default_input = Path(__file__).resolve().parent / "uploads_account"
    parser = argparse.ArgumentParser(
        description=(
            "Process Excel files one at a time through app_account.py and save "
            "the account summaries to a resumable SQLite database."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=default_input,
        help=f"Folder containing source workbooks (default: {default_input})",
    )
    parser.add_argument(
        "--database",
        type=Path,
        default=DEFAULT_DATABASE_PATH,
        help=f"SQLite database path (default: {DEFAULT_DATABASE_PATH})",
    )
    parser.add_argument(
        "--watch",
        action="store_true",
        help="Keep scanning for new files after the current queue is finished.",
    )
    parser.add_argument(
        "--scan-interval",
        type=int,
        default=60,
        help="Seconds between scans in watch mode (default: 60).",
    )
    parser.add_argument(
        "--max-attempts",
        type=int,
        default=2,
        help="Maximum attempts for a failing file (default: 2).",
    )
    parser.add_argument(
        "--max-files",
        type=int,
        help="Stop after this many files; useful for a test run.",
    )
    parser.add_argument(
        "--retry-failed",
        action="store_true",
        help="Queue all failed files again and reset their attempt count.",
    )
    parser.add_argument(
        "--reprocess-all",
        action="store_true",
        help="Queue every discovered file again without deleting old data first.",
    )
    parser.add_argument(
        "--keep-awake",
        action="store_true",
        help="Prevent Windows system sleep while the worker is running.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Write the existing app's debug output to the worker log.",
    )
    return parser


def main() -> int:
    args = build_argument_parser().parse_args()
    input_directory = args.input.expanduser().resolve()
    database_path = args.database.expanduser().resolve()
    if not input_directory.is_dir():
        print(f"Input folder does not exist: {input_directory}", file=sys.stderr)
        return 2
    if args.max_attempts < 1:
        print("--max-attempts must be at least 1.", file=sys.stderr)
        return 2

    initialize_database(database_path)
    configure_logging(database_path, args.verbose)

    try:
        with SingleWorkerLock(database_path), keep_windows_awake(args.keep_awake):
            interrupted = reset_interrupted_files(database_path)
            if interrupted:
                LOGGER.warning("Re-queued %s interrupted file(s).", interrupted)
            if args.retry_failed:
                LOGGER.info(
                    "Re-queued %s failed file(s).",
                    reset_failed_files(database_path),
                )

            discovery = discover_files(database_path, input_directory)
            LOGGER.info(
                "Scan complete: %s Excel files, %s new, %s changed.",
                discovery["seen"],
                discovery["new"],
                discovery["changed"],
            )
            if args.reprocess_all:
                LOGGER.info(
                    "Queued %s file(s) for full reprocessing.",
                    reset_all_files(database_path),
                )

            set_worker_state(
                database_path,
                is_running=True,
                process_id=os.getpid(),
                current_file=None,
                message="Worker started",
                started_at=utc_now(),
            )
            total_processed = 0
            while True:
                remaining_limit = (
                    None
                    if args.max_files is None
                    else max(0, args.max_files - total_processed)
                )
                if remaining_limit == 0:
                    break
                total_processed += process_queue(
                    database_path,
                    maximum_attempts=args.max_attempts,
                    maximum_files=remaining_limit,
                )
                if args.max_files is not None and total_processed >= args.max_files:
                    break
                if not args.watch:
                    break
                set_worker_state(
                    database_path,
                    is_running=True,
                    process_id=os.getpid(),
                    current_file=None,
                    message=f"Queue complete; next scan in {args.scan_interval}s",
                )
                LOGGER.info(
                    "Queue complete. Watching for new files every %s seconds.",
                    args.scan_interval,
                )
                time.sleep(max(5, args.scan_interval))
                discovery = discover_files(database_path, input_directory)
                if discovery["new"] or discovery["changed"]:
                    LOGGER.info(
                        "New scan: %s new, %s changed.",
                        discovery["new"],
                        discovery["changed"],
                    )

            final_counts = queue_counts(database_path, args.max_attempts)
            LOGGER.info(
                "Worker stopped cleanly. Completed: %s, failed: %s, total: %s.",
                final_counts["completed"],
                final_counts["failed"],
                final_counts["total"],
            )
            set_worker_state(
                database_path,
                is_running=False,
                process_id=None,
                current_file=None,
                message="Worker stopped cleanly",
            )
            return 0
    except KeyboardInterrupt:
        LOGGER.warning("Worker stopped by user. Progress is saved.")
        set_worker_state(
            database_path,
            is_running=False,
            process_id=None,
            current_file=None,
            message="Stopped by user; progress is saved",
        )
        return 130
    except RuntimeError as exc:
        LOGGER.error("%s", exc)
        return 1
    except Exception:
        LOGGER.exception("Worker stopped because of an unexpected error.")
        set_worker_state(
            database_path,
            is_running=False,
            process_id=None,
            current_file=None,
            message="Worker stopped after an unexpected error",
        )
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
