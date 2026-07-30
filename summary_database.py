"""SQLite storage and Excel export helpers for account summaries."""

from __future__ import annotations

import math
import sqlite3
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_DATABASE_PATH = BASE_DIR / "data" / "account_summaries.sqlite"

ACCOUNT_COLUMNS = (
    ("Acknowledgement No", "acknowledgement_no"),
    ("Bank Name", "bank_name"),
    ("Account Number", "account_number"),
    ("Credited Transaction ID", "credited_transaction_id"),
    ("Total Credited Amount", "total_credited_amount"),
    ("Total Debited Amount", "total_debited_amount"),
    ("Updated Amount (Recovery)", "updated_amount"),
    ("Not Updated Amount", "not_updated_amount"),
    ("Status", "status"),
    ("Found in Other Sheets", "found_in_other_sheets"),
    ("Breakdown by Sheet", "breakdown_by_sheet"),
    ("Duplicate Entry Info", "duplicate_entry_info"),
)

BANK_COLUMNS = (
    ("Acknowledgement No", "acknowledgement_no"),
    ("Bank Name", "bank_name"),
    ("Total Credited Amount", "total_credited_amount"),
    ("Total Debited Amount", "total_debited_amount"),
    ("Updated Amount (Recovery)", "updated_amount"),
    ("Not Updated Amount", "not_updated_amount"),
    ("Status", "status"),
    ("Found in Other Sheets", "found_in_other_sheets"),
    ("Breakdown by Sheet", "breakdown_by_sheet"),
    ("Duplicate Entry Info", "duplicate_entry_info"),
)

VIEW_CONFIG = {
    "account": {
        "table": "account_summaries",
        "sheet": "Account Wise Summary",
        "columns": ACCOUNT_COLUMNS,
        "search_columns": (
            "acknowledgement_no",
            "bank_name",
            "account_number",
            "credited_transaction_id",
        ),
        "amount_columns": {
            "total_credited_amount",
            "total_debited_amount",
            "updated_amount",
            "not_updated_amount",
        },
        "column_widths": (35, 30, 25, 24, 22, 22, 22, 22, 15, 20, 60, 60),
    },
    "bank": {
        "table": "bank_summaries",
        "sheet": "Bank Wise Summary",
        "columns": BANK_COLUMNS,
        "search_columns": ("acknowledgement_no", "bank_name"),
        "amount_columns": {
            "total_credited_amount",
            "total_debited_amount",
            "updated_amount",
            "not_updated_amount",
        },
        "column_widths": (35, 30, 22, 22, 22, 22, 15, 20, 60, 60),
    },
    "partial": {
        "table": "partial_bank_summaries",
        "sheet": "Partial Bank Wise Summary",
        "columns": BANK_COLUMNS,
        "search_columns": ("acknowledgement_no", "bank_name"),
        "amount_columns": {
            "total_credited_amount",
            "total_debited_amount",
            "updated_amount",
            "not_updated_amount",
        },
        "column_widths": (35, 30, 22, 22, 22, 22, 15, 20, 60, 60),
    },
}

SOURCE_TO_DATABASE_COLUMNS = {
    "Acknowledgement No": "acknowledgement_no",
    "Bank Name": "bank_name",
    "Account Number": "account_number",
    "Credited Transaction ID": "credited_transaction_id",
    "Total Credited Amount": "total_credited_amount",
    "Total Debited Amount": "total_debited_amount",
    "Updated Amount (Recovery)": "updated_amount",
    "Not Updated Amount": "not_updated_amount",
    "Status": "status",
    "Found in Other Sheets": "found_in_other_sheets",
    "Breakdown by Sheet": "breakdown_by_sheet",
    "Duplicate Entry Info": "duplicate_entry_info",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def connect_database(
    database_path: str | Path = DEFAULT_DATABASE_PATH,
    *,
    readonly: bool = False,
) -> sqlite3.Connection:
    path = Path(database_path).expanduser().resolve()
    if not readonly:
        path.parent.mkdir(parents=True, exist_ok=True)
        connection = sqlite3.connect(str(path), timeout=30, check_same_thread=False)
    else:
        connection = sqlite3.connect(
            f"file:{path.as_posix()}?mode=ro",
            uri=True,
            timeout=30,
            check_same_thread=False,
        )
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    connection.execute("PRAGMA busy_timeout = 30000")
    if not readonly:
        connection.execute("PRAGMA journal_mode = WAL")
        connection.execute("PRAGMA synchronous = NORMAL")
    return connection


def initialize_database(
    database_path: str | Path = DEFAULT_DATABASE_PATH,
) -> Path:
    path = Path(database_path).expanduser().resolve()
    connection = connect_database(path)
    try:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS schema_metadata (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS source_files (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_path TEXT NOT NULL COLLATE NOCASE UNIQUE,
                file_name TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                mtime_ns INTEGER NOT NULL,
                fingerprint TEXT NOT NULL,
                status TEXT NOT NULL DEFAULT 'pending'
                    CHECK (status IN ('pending', 'processing', 'completed', 'failed')),
                attempts INTEGER NOT NULL DEFAULT 0,
                discovered_at TEXT NOT NULL,
                last_seen_at TEXT NOT NULL,
                started_at TEXT,
                completed_at TEXT,
                duration_seconds REAL,
                acknowledgement_count INTEGER NOT NULL DEFAULT 0,
                account_row_count INTEGER NOT NULL DEFAULT 0,
                bank_row_count INTEGER NOT NULL DEFAULT 0,
                partial_bank_row_count INTEGER NOT NULL DEFAULT 0,
                error_message TEXT
            );

            CREATE TABLE IF NOT EXISTS account_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file_id INTEGER NOT NULL
                    REFERENCES source_files(id) ON DELETE CASCADE,
                acknowledgement_no TEXT NOT NULL COLLATE NOCASE,
                bank_name TEXT,
                account_number TEXT,
                credited_transaction_id TEXT,
                total_credited_amount REAL NOT NULL DEFAULT 0,
                total_debited_amount REAL NOT NULL DEFAULT 0,
                updated_amount REAL NOT NULL DEFAULT 0,
                not_updated_amount REAL NOT NULL DEFAULT 0,
                status TEXT,
                found_in_other_sheets TEXT,
                breakdown_by_sheet TEXT,
                duplicate_entry_info TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS bank_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file_id INTEGER NOT NULL
                    REFERENCES source_files(id) ON DELETE CASCADE,
                acknowledgement_no TEXT NOT NULL COLLATE NOCASE,
                bank_name TEXT,
                total_credited_amount REAL NOT NULL DEFAULT 0,
                total_debited_amount REAL NOT NULL DEFAULT 0,
                updated_amount REAL NOT NULL DEFAULT 0,
                not_updated_amount REAL NOT NULL DEFAULT 0,
                status TEXT,
                found_in_other_sheets TEXT,
                breakdown_by_sheet TEXT,
                duplicate_entry_info TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS partial_bank_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                source_file_id INTEGER NOT NULL
                    REFERENCES source_files(id) ON DELETE CASCADE,
                acknowledgement_no TEXT NOT NULL COLLATE NOCASE,
                bank_name TEXT,
                total_credited_amount REAL NOT NULL DEFAULT 0,
                total_debited_amount REAL NOT NULL DEFAULT 0,
                updated_amount REAL NOT NULL DEFAULT 0,
                not_updated_amount REAL NOT NULL DEFAULT 0,
                status TEXT,
                found_in_other_sheets TEXT,
                breakdown_by_sheet TEXT,
                duplicate_entry_info TEXT,
                created_at TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS worker_state (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                is_running INTEGER NOT NULL DEFAULT 0,
                process_id INTEGER,
                started_at TEXT,
                heartbeat_at TEXT,
                current_file TEXT,
                message TEXT
            );

            CREATE INDEX IF NOT EXISTS idx_source_files_status
                ON source_files(status, attempts, id);
            CREATE INDEX IF NOT EXISTS idx_account_ack
                ON account_summaries(acknowledgement_no);
            CREATE INDEX IF NOT EXISTS idx_account_status
                ON account_summaries(status);
            CREATE INDEX IF NOT EXISTS idx_account_bank
                ON account_summaries(bank_name);
            CREATE INDEX IF NOT EXISTS idx_account_number
                ON account_summaries(account_number);
            CREATE INDEX IF NOT EXISTS idx_bank_ack
                ON bank_summaries(acknowledgement_no);
            CREATE INDEX IF NOT EXISTS idx_bank_status
                ON bank_summaries(status);
            CREATE INDEX IF NOT EXISTS idx_partial_ack
                ON partial_bank_summaries(acknowledgement_no);

            INSERT INTO schema_metadata(key, value)
            VALUES ('schema_version', '1')
            ON CONFLICT(key) DO UPDATE SET value = excluded.value;

            INSERT OR IGNORE INTO worker_state(id, is_running)
            VALUES (1, 0);
            """
        )
        connection.commit()
    finally:
        connection.close()
    return path


def _clean_text(value: Any, default: str = "") -> str:
    if value is None:
        return default
    if isinstance(value, float) and math.isnan(value):
        return default
    text = str(value).strip()
    if text.lower() in {"nan", "none"}:
        return default
    return text


def _clean_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, float) and math.isnan(value):
        return 0.0
    if isinstance(value, str):
        value = (
            value.replace(",", "")
            .replace("\u20b9", "")
            .replace("â‚¹", "")
            .strip()
        )
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _record_to_values(
    record: Mapping[str, Any],
    columns: Sequence[tuple[str, str]],
) -> tuple[Any, ...]:
    values: list[Any] = []
    amount_fields = {
        "total_credited_amount",
        "total_debited_amount",
        "updated_amount",
        "not_updated_amount",
    }
    for display_name, database_name in columns:
        value = record.get(display_name)
        if value is None:
            value = record.get(database_name)
        if database_name in amount_fields:
            values.append(_clean_number(value))
        else:
            values.append(_clean_text(value))
    return tuple(values)


def save_file_summaries(
    database_path: str | Path,
    source_file_id: int,
    summaries: Mapping[str, Sequence[Mapping[str, Any]]],
    *,
    duration_seconds: float,
) -> dict[str, int]:
    """Atomically replace one source file's summary rows and mark it completed."""
    account_records = list(summaries.get("Account Wise Summary", ()))
    bank_records = list(summaries.get("Bank Wise Summary", ()))
    partial_records = list(summaries.get("Partial Bank Wise Summary", ()))
    created_at = utc_now()

    account_values = [
        (source_file_id, *_record_to_values(record, ACCOUNT_COLUMNS), created_at)
        for record in account_records
    ]
    bank_values = [
        (source_file_id, *_record_to_values(record, BANK_COLUMNS), created_at)
        for record in bank_records
    ]
    partial_values = [
        (source_file_id, *_record_to_values(record, BANK_COLUMNS), created_at)
        for record in partial_records
    ]
    acknowledgements = {
        values[1]
        for values in account_values
        if values[1]
    }

    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute(
                "DELETE FROM account_summaries WHERE source_file_id = ?",
                (source_file_id,),
            )
            connection.execute(
                "DELETE FROM bank_summaries WHERE source_file_id = ?",
                (source_file_id,),
            )
            connection.execute(
                "DELETE FROM partial_bank_summaries WHERE source_file_id = ?",
                (source_file_id,),
            )

            if account_values:
                connection.executemany(
                    """
                    INSERT INTO account_summaries (
                        source_file_id,
                        acknowledgement_no,
                        bank_name,
                        account_number,
                        credited_transaction_id,
                        total_credited_amount,
                        total_debited_amount,
                        updated_amount,
                        not_updated_amount,
                        status,
                        found_in_other_sheets,
                        breakdown_by_sheet,
                        duplicate_entry_info,
                        created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    account_values,
                )
            if bank_values:
                connection.executemany(
                    """
                    INSERT INTO bank_summaries (
                        source_file_id,
                        acknowledgement_no,
                        bank_name,
                        total_credited_amount,
                        total_debited_amount,
                        updated_amount,
                        not_updated_amount,
                        status,
                        found_in_other_sheets,
                        breakdown_by_sheet,
                        duplicate_entry_info,
                        created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    bank_values,
                )
            if partial_values:
                connection.executemany(
                    """
                    INSERT INTO partial_bank_summaries (
                        source_file_id,
                        acknowledgement_no,
                        bank_name,
                        total_credited_amount,
                        total_debited_amount,
                        updated_amount,
                        not_updated_amount,
                        status,
                        found_in_other_sheets,
                        breakdown_by_sheet,
                        duplicate_entry_info,
                        created_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    partial_values,
                )

            connection.execute(
                """
                UPDATE source_files
                SET status = 'completed',
                    completed_at = ?,
                    duration_seconds = ?,
                    acknowledgement_count = ?,
                    account_row_count = ?,
                    bank_row_count = ?,
                    partial_bank_row_count = ?,
                    error_message = NULL
                WHERE id = ?
                """,
                (
                    created_at,
                    duration_seconds,
                    len(acknowledgements),
                    len(account_values),
                    len(bank_values),
                    len(partial_values),
                    source_file_id,
                ),
            )
        return {
            "acknowledgements": len(acknowledgements),
            "account": len(account_values),
            "bank": len(bank_values),
            "partial": len(partial_values),
        }
    finally:
        connection.close()


def mark_file_failed(
    database_path: str | Path,
    source_file_id: int,
    error_message: str,
    *,
    duration_seconds: float,
) -> None:
    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute(
                """
                UPDATE source_files
                SET status = 'failed',
                    completed_at = ?,
                    duration_seconds = ?,
                    error_message = ?
                WHERE id = ?
                """,
                (
                    utc_now(),
                    duration_seconds,
                    error_message[:4000],
                    source_file_id,
                ),
            )
    finally:
        connection.close()


def set_worker_state(
    database_path: str | Path,
    *,
    is_running: bool,
    process_id: int | None = None,
    current_file: str | None = None,
    message: str | None = None,
    started_at: str | None = None,
) -> None:
    now = utc_now()
    connection = connect_database(database_path)
    try:
        with connection:
            connection.execute(
                """
                UPDATE worker_state
                SET is_running = ?,
                    process_id = ?,
                    started_at = COALESCE(?, started_at),
                    heartbeat_at = ?,
                    current_file = ?,
                    message = ?
                WHERE id = 1
                """,
                (
                    int(is_running),
                    process_id,
                    started_at,
                    now,
                    current_file,
                    message,
                ),
            )
    finally:
        connection.close()


def query_progress(database_path: str | Path) -> dict[str, Any]:
    initialize_database(database_path)
    connection = connect_database(database_path, readonly=True)
    try:
        status_rows = connection.execute(
            """
            SELECT status, COUNT(*) AS file_count
            FROM source_files
            GROUP BY status
            """
        ).fetchall()
        counts = {
            "pending": 0,
            "processing": 0,
            "completed": 0,
            "failed": 0,
        }
        counts.update({row["status"]: row["file_count"] for row in status_rows})
        total = sum(counts.values())
        processed = counts["completed"] + counts["failed"]
        percent = round((processed / total * 100), 2) if total else 0.0
        totals = connection.execute(
            """
            SELECT
                COALESCE(SUM(acknowledgement_count), 0) AS acknowledgements,
                COALESCE(SUM(account_row_count), 0) AS account_rows,
                COALESCE(SUM(bank_row_count), 0) AS bank_rows,
                COALESCE(SUM(partial_bank_row_count), 0) AS partial_rows
            FROM source_files
            WHERE status = 'completed'
            """
        ).fetchone()
        worker = connection.execute(
            "SELECT * FROM worker_state WHERE id = 1"
        ).fetchone()
        return {
            "files": {
                "total": total,
                **counts,
                "processed": processed,
                "percent": percent,
            },
            "summaries": dict(totals),
            "worker": dict(worker) if worker else {},
        }
    finally:
        connection.close()


def query_recent_failures(
    database_path: str | Path,
    *,
    limit: int = 10,
) -> list[dict[str, Any]]:
    initialize_database(database_path)
    connection = connect_database(database_path, readonly=True)
    try:
        rows = connection.execute(
            """
            SELECT file_name, source_path, attempts, completed_at, error_message
            FROM source_files
            WHERE status = 'failed'
            ORDER BY completed_at DESC, id DESC
            LIMIT ?
            """,
            (max(1, min(limit, 100)),),
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()


def _build_filter_clause(
    config: Mapping[str, Any],
    *,
    acknowledgement: str | None,
    status: str | None,
    search: str | None,
) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    parameters: list[Any] = []
    if acknowledgement and acknowledgement.upper() != "ALL":
        clauses.append("acknowledgement_no = ? COLLATE NOCASE")
        parameters.append(acknowledgement.strip())
    if status and status.upper() not in {"", "ALL"}:
        clauses.append("UPPER(status) = ?")
        parameters.append(status.upper())
    if search:
        search_term = f"%{search.strip()}%"
        search_clauses = []
        for column in config["search_columns"]:
            search_clauses.append(f"COALESCE({column}, '') LIKE ?")
            parameters.append(search_term)
        clauses.append(f"({' OR '.join(search_clauses)})")
    if not clauses:
        return "", parameters
    return " WHERE " + " AND ".join(clauses), parameters


def query_summary_page(
    database_path: str | Path,
    *,
    view: str = "account",
    acknowledgement: str | None = "ALL",
    status: str | None = "ALL",
    search: str | None = None,
    page: int = 1,
    page_size: int = 100,
) -> dict[str, Any]:
    initialize_database(database_path)
    config = VIEW_CONFIG.get(view, VIEW_CONFIG["account"])
    page = max(1, page)
    page_size = max(10, min(page_size, 500))
    where_clause, parameters = _build_filter_clause(
        config,
        acknowledgement=acknowledgement,
        status=status,
        search=search,
    )
    database_columns = [database_name for _, database_name in config["columns"]]
    select_columns = ", ".join(database_columns)

    connection = connect_database(database_path, readonly=True)
    try:
        total = connection.execute(
            f"SELECT COUNT(*) FROM {config['table']}{where_clause}",
            parameters,
        ).fetchone()[0]
        page_count = max(1, math.ceil(total / page_size))
        page = min(page, page_count)
        rows = connection.execute(
            f"""
            SELECT {select_columns}
            FROM {config['table']}
            {where_clause}
            ORDER BY not_updated_amount DESC, acknowledgement_no, id
            LIMIT ? OFFSET ?
            """,
            [*parameters, page_size, (page - 1) * page_size],
        ).fetchall()
        return {
            "view": view,
            "columns": [
                {
                    "label": display_name,
                    "key": database_name,
                    "type": (
                        "amount"
                        if database_name in config["amount_columns"]
                        else (
                            "status"
                            if database_name == "status"
                            else "text"
                        )
                    ),
                }
                for display_name, database_name in config["columns"]
            ],
            "rows": [dict(row) for row in rows],
            "pagination": {
                "page": page,
                "page_size": page_size,
                "page_count": page_count,
                "total": total,
            },
        }
    finally:
        connection.close()


def query_acknowledgements(
    database_path: str | Path,
    *,
    search: str | None = None,
    limit: int = 50,
) -> list[dict[str, Any]]:
    initialize_database(database_path)
    clauses = ""
    parameters: list[Any] = []
    if search:
        clauses = "WHERE acknowledgement_no LIKE ?"
        parameters.append(f"%{search.strip()}%")
    parameters.append(max(1, min(limit, 500)))
    connection = connect_database(database_path, readonly=True)
    try:
        rows = connection.execute(
            f"""
            SELECT acknowledgement_no, COUNT(*) AS account_rows
            FROM account_summaries
            {clauses}
            GROUP BY acknowledgement_no
            ORDER BY acknowledgement_no
            LIMIT ?
            """,
            parameters,
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        connection.close()


def query_summary_totals(
    database_path: str | Path,
    *,
    acknowledgement: str | None = "ALL",
) -> dict[str, Any]:
    initialize_database(database_path)
    where_clause = ""
    parameters: list[Any] = []
    if acknowledgement and acknowledgement.upper() != "ALL":
        where_clause = "WHERE acknowledgement_no = ? COLLATE NOCASE"
        parameters.append(acknowledgement.strip())
    connection = connect_database(database_path, readonly=True)
    try:
        row = connection.execute(
            f"""
            SELECT
                COUNT(*) AS account_rows,
                COUNT(DISTINCT acknowledgement_no) AS acknowledgements,
                COALESCE(SUM(total_credited_amount), 0) AS total_credited,
                COALESCE(SUM(total_debited_amount), 0) AS total_debited,
                COALESCE(SUM(updated_amount), 0) AS total_updated,
                COALESCE(SUM(not_updated_amount), 0) AS total_not_updated,
                SUM(CASE WHEN UPPER(status) = 'PENDING' THEN 1 ELSE 0 END)
                    AS pending_rows,
                SUM(CASE WHEN UPPER(status) = 'PARTIAL' THEN 1 ELSE 0 END)
                    AS partial_rows,
                SUM(CASE WHEN UPPER(status) IN ('COMPLETED', 'COMPLETE')
                    THEN 1 ELSE 0 END) AS completed_rows
            FROM account_summaries
            {where_clause}
            """,
            parameters,
        ).fetchone()
        result = dict(row)
        for key in ("pending_rows", "partial_rows", "completed_rows"):
            result[key] = result[key] or 0
        return result
    finally:
        connection.close()


def _iter_export_rows(
    connection: sqlite3.Connection,
    config: Mapping[str, Any],
    acknowledgement: str | None,
) -> Iterable[sqlite3.Row]:
    where_clause = ""
    parameters: list[Any] = []
    if acknowledgement and acknowledgement.upper() != "ALL":
        where_clause = "WHERE acknowledgement_no = ? COLLATE NOCASE"
        parameters.append(acknowledgement.strip())
    database_columns = ", ".join(
        database_name for _, database_name in config["columns"]
    )
    cursor = connection.execute(
        f"""
        SELECT {database_columns}
        FROM {config['table']}
        {where_clause}
        ORDER BY not_updated_amount DESC, acknowledgement_no, id
        """,
        parameters,
    )
    while True:
        rows = cursor.fetchmany(1000)
        if not rows:
            break
        yield from rows


def create_excel_export(
    database_path: str | Path,
    *,
    acknowledgement: str | None = "ALL",
) -> BytesIO:
    """Build the same three summary views as a styled, filterable workbook."""
    initialize_database(database_path)
    connection = connect_database(database_path, readonly=True)
    workbook = Workbook(write_only=True)

    header_fill = PatternFill(
        start_color="1F2937",
        end_color="1F2937",
        fill_type="solid",
    )
    header_font = Font(color="FFFFFF", bold=True, size=11)
    stripe_fill = PatternFill(
        start_color="F8FAFC",
        end_color="F8FAFC",
        fill_type="solid",
    )
    status_styles = {
        "PENDING": (
            PatternFill("solid", fgColor="FEE2E2"),
            Font(color="B91C1C", bold=True),
        ),
        "PARTIAL": (
            PatternFill("solid", fgColor="FEF3C7"),
            Font(color="B45309", bold=True),
        ),
        "COMPLETED": (
            PatternFill("solid", fgColor="DCFCE7"),
            Font(color="15803D", bold=True),
        ),
        "COMPLETE": (
            PatternFill("solid", fgColor="DCFCE7"),
            Font(color="15803D", bold=True),
        ),
    }
    thin_border = Border(
        bottom=Side(style="hair", color="CBD5E1"),
    )

    try:
        connection.execute("BEGIN")
        for view_name in ("account", "bank", "partial"):
            config = VIEW_CONFIG[view_name]
            worksheet = workbook.create_sheet(config["sheet"])
            worksheet.freeze_panes = "A2"
            worksheet.sheet_view.showGridLines = False

            header_cells = []
            for display_name, _ in config["columns"]:
                cell = WriteOnlyCell(worksheet, value=display_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                header_cells.append(cell)
            worksheet.append(header_cells)

            row_count = 0
            amount_columns = config["amount_columns"]
            for row_count, row in enumerate(
                _iter_export_rows(connection, config, acknowledgement),
                start=1,
            ):
                excel_cells = []
                striped = row_count % 2 == 0
                for _, database_name in config["columns"]:
                    value = row[database_name]
                    cell = WriteOnlyCell(worksheet, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if striped:
                        cell.fill = stripe_fill
                    if database_name in amount_columns:
                        cell.number_format = "\u20b9#,##0.00"
                        cell.alignment = Alignment(
                            horizontal="right",
                            vertical="center",
                        )
                    elif database_name == "status":
                        status_value = _clean_text(value).upper()
                        style = status_styles.get(status_value)
                        if style:
                            cell.fill, cell.font = style
                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                        )
                    elif database_name == "found_in_other_sheets":
                        cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center",
                        )
                    excel_cells.append(cell)
                worksheet.append(excel_cells)

            last_column = get_column_letter(len(config["columns"]))
            worksheet.auto_filter.ref = f"A1:{last_column}{row_count + 1}"
            for index, width in enumerate(config["column_widths"], start=1):
                worksheet.column_dimensions[get_column_letter(index)].width = width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        connection.rollback()
        return output
    finally:
        connection.close()
