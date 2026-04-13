from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import os
from datetime import datetime
from io import BytesIO

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Global data storage
df_main = None
df_other_sheets = {}
uploaded_files_count = 0

# Lookup maps for optimization
debited_trans_id_map = {}
credited_trans_id_map = {}
breakdown_map = {}

def rebuild_maps():
    """Rebuild lookup maps for faster processing"""
    global debited_trans_id_map, credited_trans_id_map, breakdown_map
    
    # Rebuild parent-child maps for df_main
    debited_trans_id_map = {}
    credited_trans_id_map = {}
    
    if df_main is not None:
        for idx, row in df_main.iterrows():
            deb = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
            cre = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
            
            if deb:
                if deb not in debited_trans_id_map:
                    debited_trans_id_map[deb] = []
                debited_trans_id_map[deb].append(idx)
            
            if cre:
                if cre not in credited_trans_id_map:
                    credited_trans_id_map[cre] = []
                credited_trans_id_map[cre].append(idx)
    
    # Rebuild breakdown map for other sheets
    breakdown_map = {}
    if df_other_sheets:
        for sheet_name, sheet_info in df_other_sheets.items():
            df = sheet_info['data']
            amount_col = sheet_info['amount_col']
            
            if len(df.columns) > 3:
                for _, row in df.iterrows():
                    trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                    if trans_id:
                        if trans_id not in breakdown_map:
                            breakdown_map[trans_id] = []
                        
                        if len(row) > amount_col:
                            amount = clean_amount(row.iloc[amount_col])
                            if amount > 0:
                                breakdown_map[trans_id].append({
                                    'sheet': sheet_name,
                                    'amount': amount
                                })

def clean_amount(value):
    """Convert amount to float"""
    if pd.isna(value) or value == '':
        return 0.0
    if isinstance(value, str):
        value = value.replace(',', '').replace('₹', '').strip()
    try:
        return float(value)
    except:
        return 0.0

def process_excel_file(filepath, is_first_file=False):
    """Process Excel file and merge with existing data"""
    global df_main, df_other_sheets, uploaded_files_count
    
    try:
        xl = pd.ExcelFile(filepath)
        
        print(f"DEBUG: Processing file with sheets: {xl.sheet_names}")
        
        # Find Money Transfer sheet
        money_transfer_sheet = None
        for sheet in xl.sheet_names:
            print(f"DEBUG: Checking sheet '{sheet}', lowercase: '{sheet.lower()}'")
            if 'money transfer' in sheet.lower():
                money_transfer_sheet = sheet
                print(f"DEBUG: Found Money Transfer sheet: '{money_transfer_sheet}'")
                break
        
        if not money_transfer_sheet:
            print(f"DEBUG: Money Transfer sheet NOT FOUND in sheets: {xl.sheet_names}")
            return False, f"Money Transfer sheet not found. Available sheets: {', '.join(xl.sheet_names)}"
        
        # Load main sheet
        df_new_main = pd.read_excel(xl, sheet_name=money_transfer_sheet)
        print(f"DEBUG: Loaded main sheet with {len(df_new_main)} rows")
        
        # Deduplicate only perfectly identical rows to avoid losing unique debited info
        df_new_main = df_new_main.drop_duplicates(keep='first')
        print(f"DEBUG: After deduplication, main sheet has {len(df_new_main)} rows")
        
        # Merge or initialize main dataframe
        if is_first_file or df_main is None:
            df_main = df_new_main
            df_other_sheets = {}
        else:
            df_main = pd.concat([df_main, df_new_main], ignore_index=True)
            # Deduplicate again after merge to remove perfectly identical rows
            df_main = df_main.drop_duplicates(keep='first')
            
        # Reset index to ensure it's contiguous after dropping duplicates
        df_main = df_main.reset_index(drop=True)
        
        # Load and merge other sheets
        for sheet in xl.sheet_names:
            if sheet == money_transfer_sheet:
                continue
            
            df_sheet = pd.read_excel(xl, sheet_name=sheet)
            
            # Determine amount column based on sheet name
            sheet_lower = sheet.lower()
            if 'withdrawal through atm' in sheet_lower:
                amount_col_idx = 6  # Column G
            elif 'cash withdrawal through cheque' in sheet_lower:
                amount_col_idx = 9  # Column J
            elif 'withdrawal through pos' in sheet_lower:
                amount_col_idx = 6  # Column G
            elif 'cheque' in sheet_lower:
                amount_col_idx = 8  # Column I (for other cheque sheets)
            else:
                amount_col_idx = 5  # Column F (default)
            
            # Merge with existing sheet data or create new
            if sheet in df_other_sheets:
                df_other_sheets[sheet]['data'] = pd.concat([df_other_sheets[sheet]['data'], df_sheet], ignore_index=True)
            else:
                df_other_sheets[sheet] = {
                    'data': df_sheet,
                    'amount_col': amount_col_idx
                }
        
        uploaded_files_count += 1
        rebuild_maps()
        return True, f"File processed successfully. Total files: {uploaded_files_count}"
        
    except Exception as e:
        return False, f"Error: {str(e)}"

def get_transaction_breakdown(trans_id):
    """Get breakdown from other sheets - ONLY NON-ZERO amounts - Optimized with breakdown_map"""
    return breakdown_map.get(str(trans_id).strip(), [])

def calculate_status(child_trans_id, disputed_amount, current_layer):
    """Calculate transaction status - accounts for BOTH children flow AND other-sheet recovery"""
    if df_main is None:
        return {
            'status': 'PENDING',
            'updated_amount': 0,
            'pending_amount': disputed_amount,
            'has_children': False,
            'in_other_sheets': False,
            'children_total': 0
        }
    
    # Optimized child lookup: Debited Trans ID (col D, index 3) matches this Child Trans ID
    # AND layer is higher than current
    has_children = False
    children_total = 0
    child_id_str = str(child_trans_id).strip()
    if child_id_str in debited_trans_id_map:
        for idx in debited_trans_id_map[child_id_str]:
            if df_main.iloc[idx, 5] > current_layer:
                has_children = True
                children_total += clean_amount(df_main.iloc[idx, 11])
    
    # Get breakdown from breakdown_map (only non-zero)
    breakdown = get_transaction_breakdown(child_trans_id)
    updated_amount = sum(item['amount'] for item in breakdown)
    in_other_sheets = len(breakdown) > 0
    
    # Total accounted = money that flowed to children + money recovered via other sheets
    total_accounted = children_total + updated_amount
    pending_amount = disputed_amount - total_accounted
    
    # Status logic
    if not has_children and not in_other_sheets:
        status = 'PENDING'
    elif pending_amount > 0.01:  # Allow small rounding errors
        status = 'PARTIAL'
    else:
        status = 'COMPLETE'
    
    return {
        'status': status,
        'updated_amount': updated_amount,
        'pending_amount': max(0, pending_amount),
        'has_children': has_children,
        'in_other_sheets': in_other_sheets,
        'children_total': children_total
    }

def get_layer_transactions(layer=1, parent_trans_id=None):
    """Get transactions for a layer"""
    if df_main is None:
        return []
    
    try:
        if layer == 1 and parent_trans_id is None:
            # Root level: start from minimum layer in dataset
            # Filter out NaN values first
            valid_layers = df_main[pd.notna(df_main.iloc[:, 5])]
            if len(valid_layers) == 0:
                print("ERROR: No valid layer values found")
                return []
            
            min_layer = int(valid_layers.iloc[:, 5].min())
            print(f"DEBUG: Minimum layer found: {min_layer}")
            print(f"DEBUG: Total rows in df_main: {len(df_main)}")
            filtered = df_main[df_main.iloc[:, 5] == min_layer]
            print(f"DEBUG: Filtered rows for layer {min_layer}: {len(filtered)}")
        else:
            # Optimized parent layer lookup
            parent_id_str = str(parent_trans_id).strip()
            if parent_id_str not in credited_trans_id_map:
                return []
            
            parent_idx = credited_trans_id_map[parent_id_str][0]
            parent_layer = int(df_main.iloc[parent_idx, 5])
            
            # Children: Parent Trans ID (col D) = parent_trans_id AND layer > parent_layer
            indices = debited_trans_id_map.get(parent_id_str, [])
            filtered_indices = [idx for idx in indices if df_main.iloc[idx, 5] > parent_layer]
            filtered = df_main.iloc[filtered_indices]
    except Exception as e:
        print(f"ERROR in get_layer_transactions: {e}")
        return []
    
    transactions = []
    for _, row in filtered.iterrows():
        child_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
        disputed_amount = clean_amount(row.iloc[11])
        current_layer = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 1
        
        status_info = calculate_status(child_trans_id, disputed_amount, current_layer)
        
        # Override status to match Excel logic: if has children = "TRANSACTION CONTINUE", else use calculated status
        if status_info['has_children']:
            final_status = 'TRANSACTION CONTINUE'
        else:
            final_status = status_info['status']
        
        # Get child bank (Optimized)
        child_bank = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ''
        child_id_str = str(child_trans_id).strip()
        if child_id_str in debited_trans_id_map:
            for c_idx in debited_trans_id_map[child_id_str]:
                if df_main.iloc[c_idx, 5] > current_layer:
                    child_bank = str(df_main.iloc[c_idx, 4])
                    break
        
        transactions.append({
            's_no': int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
            'acknowledgement_no': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
            'debited_account': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
            'debited_transaction_id': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
            'bank': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
            'layer': current_layer,
            'credited_account': str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
            'ifsc_code': str(row.iloc[7]) if pd.notna(row.iloc[7]) else '',
            'transaction_date': str(row.iloc[8]) if pd.notna(row.iloc[8]) else '',
            'credited_transaction_id': child_trans_id,
            'transaction_amount': clean_amount(row.iloc[10]),
            'disputed_amount': disputed_amount,
            'reference_no': str(row.iloc[12]) if pd.notna(row.iloc[12]) else '',
            'remarks': str(row.iloc[13]) if pd.notna(row.iloc[13]) else '',
            'action_taken_by': str(row.iloc[14]) if pd.notna(row.iloc[14]) else '',
            'date_of_action': str(row.iloc[15]) if pd.notna(row.iloc[15]) else '',
            'credited_bank': child_bank,
            'status': final_status,
            'updated_amount': status_info['updated_amount'],
            'pending_amount': status_info['pending_amount'],
            'has_children': status_info['has_children'],
            'in_other_sheets': status_info['in_other_sheets']
        })
    
    return transactions

def build_hierarchical_data(parent_id=None, layer=1, parent_path="", level=0, visited=None):
    """Build hierarchical data for Excel - INCLUDES ALL TRANSACTIONS"""
    if df_main is None:
        return []
    
    # SIMPLE APPROACH: Process EVERY row exactly once, track by row index not trans ID
    if parent_id is None and layer == 1:
        all_data = []
        processed_row_indices = set()
        
        # Build parent-child map: Key = Credited Trans ID, Value = list of row indices
        parent_child_map = {}
        for idx, row in df_main.iterrows():
            debited_trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
            if debited_trans_id:
                if debited_trans_id not in parent_child_map:
                    parent_child_map[debited_trans_id] = []
                parent_child_map[debited_trans_id].append(idx)
        
        # Find root rows: Credited Trans ID NOT in any Debited Trans ID
        all_credited_ids = set(str(row.iloc[9]).strip() for _, row in df_main.iterrows() if pd.notna(row.iloc[9]) and str(row.iloc[9]).strip())
        all_debited_ids = set(str(row.iloc[3]).strip() for _, row in df_main.iterrows() if pd.notna(row.iloc[3]) and str(row.iloc[3]).strip())
        root_credited_ids = all_credited_ids - all_debited_ids
        
        # Process root rows
        for idx, row in df_main.iterrows():
            if idx in processed_row_indices:
                continue
            credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
            if credited_trans_id in root_credited_ids:
                processed_row_indices.add(idx)
                # Root path is just the credited trans ID
                row_data = process_single_row(row, idx, level=0, parent_path=credited_trans_id, parent_child_map=parent_child_map)
                all_data.append(row_data)
                
                # Add children recursively
                children_data = process_children_by_index(credited_trans_id, level=1, parent_path=credited_trans_id, 
                                                         parent_child_map=parent_child_map, processed_row_indices=processed_row_indices)
                all_data.extend(children_data)
        
        # Process remaining rows (orphans, circular refs, or middle-of-chain entries)
        for idx, row in df_main.iterrows():
            if idx not in processed_row_indices:
                processed_row_indices.add(idx)
                credited_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
                debited_trans_id = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else ''
                
                # Build path: show both debited and credited trans IDs
                if debited_trans_id and credited_trans_id:
                    path = f"{debited_trans_id} → {credited_trans_id}"
                elif credited_trans_id:
                    path = credited_trans_id
                else:
                    path = debited_trans_id if debited_trans_id else "No Trans ID"
                
                row_data = process_single_row(row, idx, level=0, parent_path=path, parent_child_map=parent_child_map)
                all_data.append(row_data)
                
                # Try to add children
                if credited_trans_id:
                    children_data = process_children_by_index(credited_trans_id, level=1, parent_path=path, 
                                                             parent_child_map=parent_child_map, processed_row_indices=processed_row_indices)
                    all_data.extend(children_data)
        
        print(f"DEBUG: Processed {len(processed_row_indices)} rows out of {len(df_main)} total rows")
        print(f"DEBUG: Output has {len(all_data)} rows")
        
        return all_data
    
    return []

def process_single_row(row, row_idx, level, parent_path, parent_child_map):
    """Process a single row and return formatted data"""
    child_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
    disputed_amount = clean_amount(row.iloc[11])
    current_layer = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 1
    
    status_info = calculate_status(child_trans_id, disputed_amount, current_layer)
    
    # Override status: if has children = "TRANSACTION CONTINUE", else use calculated status
    if status_info['has_children']:
        final_status = 'TRANSACTION CONTINUE'
    else:
        final_status = status_info['status']
    
    # Calculate children amount difference (using pre-computed children_total from calculate_status)
    if status_info['has_children']:
        amount_difference = disputed_amount - status_info['children_total']
    else:
        amount_difference = None
    
    breakdown = get_transaction_breakdown(child_trans_id)
    breakdown_text = "; ".join([f"{item['sheet']}: ₹{item['amount']:,.2f}" for item in breakdown]) if breakdown else "None"
    
    child_bank = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ''
    child_id_str = str(child_trans_id).strip()
    if child_id_str in debited_trans_id_map:
        for c_idx in debited_trans_id_map[child_id_str]:
            if df_main.iloc[c_idx, 5] > current_layer:
                child_bank = str(df_main.iloc[c_idx, 4])
                break
    
    # Use the parent_path as-is (don't modify it here)
    current_path = parent_path
    
    row_data = {
        'Level': level,
        'Layer': current_layer,
        'Debited Trans ID (Filter)': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
        'Hierarchy Path': current_path,
        'S.No': int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
        'Debited Account': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
        'Debited Transaction ID': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
        'Debited Bank': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
        'Acknowledgement No': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
        'Credited Account': str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
        'Credited Transaction ID': child_trans_id,
        'Credited Bank': child_bank,
        'IFSC Code': str(row.iloc[7]) if pd.notna(row.iloc[7]) else '',
        'Transaction Date': str(row.iloc[8]) if pd.notna(row.iloc[8]) else '',
        'Transaction Amount': clean_amount(row.iloc[10]),
        'Disputed Amount': disputed_amount,
        'Layerwise Amount Difference': amount_difference if amount_difference is not None else '',
        'Updated Amount': status_info['updated_amount'],
        'Pending Amount': status_info['pending_amount'],
        'Status': final_status,
        'Has Children': 'Yes' if status_info['has_children'] else 'No',
        'Found in Other Sheets': 'Yes' if status_info['in_other_sheets'] else 'No',
        'Breakdown by Sheet': breakdown_text,
        'Reference No': str(row.iloc[12]) if pd.notna(row.iloc[12]) else '',
        'Remarks': str(row.iloc[13]) if pd.notna(row.iloc[13]) else '',
        'Action Taken By': str(row.iloc[14]) if pd.notna(row.iloc[14]) else '',
        'Date of Action': str(row.iloc[15]) if pd.notna(row.iloc[15]) else ''
    }
    
    return row_data

def process_children_by_index(parent_trans_id, level, parent_path, parent_child_map, processed_row_indices, max_depth=20):
    """Recursively process children by row index to avoid duplicates"""
    if level > max_depth:
        return []
    
    children_data = []
    
    if parent_trans_id in parent_child_map:
        for child_idx in parent_child_map[parent_trans_id]:
            if child_idx in processed_row_indices:
                continue
            
            processed_row_indices.add(child_idx)
            child_row = df_main.iloc[child_idx]
            child_credited_trans_id = str(child_row.iloc[9]).strip() if pd.notna(child_row.iloc[9]) else ''
            
            # Build the child's path by appending to parent path
            child_path = f"{parent_path} → {child_credited_trans_id}" if child_credited_trans_id else parent_path
            
            # Pass the child's path (not parent's path)
            row_data = process_single_row(child_row, child_idx, level=level, parent_path=child_path, parent_child_map=parent_child_map)
            children_data.append(row_data)
            
            # Recursively process grandchildren with child's path
            grandchildren_data = process_children_by_index(child_credited_trans_id, level=level+1, 
                                                          parent_path=child_path,
                                                          parent_child_map=parent_child_map, 
                                                          processed_row_indices=processed_row_indices,
                                                          max_depth=max_depth)
            children_data.extend(grandchildren_data)
    
    return children_data

@app.route('/')
def index():
    response = render_template('index.html')
    return response, 200, {
        'Cache-Control': 'no-cache, no-store, must-revalidate',
        'Pragma': 'no-cache',
        'Expires': '0'
    }

@app.route('/test')
def test():
    return render_template('test.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    global uploaded_files_count
    
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    files = request.files.getlist('file')
    
    if not files or files[0].filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if len(files) > 100:
        return jsonify({'success': False, 'message': 'Maximum 100 files allowed'})
    
    # Check if this is a fresh upload (clear previous data)
    clear_data = request.form.get('clear_data', 'true').lower() == 'true'
    
    if clear_data:
        uploaded_files_count = 0
    
    success_count = 0
    error_messages = []
    
    for idx, file in enumerate(files):
        if not file.filename.endswith(('.xlsx', '.xls')):
            error_messages.append(f"{file.filename}: Only Excel files allowed")
            continue
        
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'file_{idx}_{file.filename}')
        file.save(filepath)
        
        is_first = (idx == 0 and clear_data)
        success, message = process_excel_file(filepath, is_first_file=is_first)
        
        if success:
            success_count += 1
        else:
            error_messages.append(f"{file.filename}: {message}")
    
    if success_count > 0:
        msg = f"✓ Successfully processed {success_count} file(s). Total files loaded: {uploaded_files_count}"
        if error_messages:
            msg += f"\n⚠ Errors: {'; '.join(error_messages)}"
        return jsonify({'success': True, 'message': msg})
    else:
        return jsonify({'success': False, 'message': f"Failed to process files. Errors: {'; '.join(error_messages)}"})

@app.route('/clear-data', methods=['POST'])
def clear_data():
    """Clear all uploaded data"""
    global df_main, df_other_sheets, uploaded_files_count
    df_main = None
    df_other_sheets = {}
    uploaded_files_count = 0
    rebuild_maps()
    return jsonify({'success': True, 'message': 'All data cleared'})

@app.route('/api/min-layer')
def get_min_layer():
    """Get the minimum layer in the dataset"""
    if df_main is None:
        return jsonify({'min_layer': 1})
    
    try:
        valid_layers = df_main[pd.notna(df_main.iloc[:, 5])]
        if len(valid_layers) == 0:
            return jsonify({'min_layer': 1})
        min_layer = int(valid_layers.iloc[:, 5].min())
        return jsonify({'min_layer': min_layer})
    except:
        return jsonify({'min_layer': 1})

@app.route('/api/transactions')
def get_transactions():
    layer = int(request.args.get('layer', 1))
    parent_id = request.args.get('parent_id', None)
    
    transactions = get_layer_transactions(layer, parent_id)
    return jsonify(transactions)

@app.route('/api/all-transactions')
def get_all_transactions():
    """Get ALL transactions from the dataset for filtering"""
    if df_main is None:
        return jsonify([])
    
    all_transactions = []
    
    for _, row in df_main.iterrows():
        child_trans_id = str(row.iloc[9]).strip() if pd.notna(row.iloc[9]) else ''
        disputed_amount = clean_amount(row.iloc[11])
        current_layer = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 1
        
        status_info = calculate_status(child_trans_id, disputed_amount, current_layer)
        
        # Override status to match Excel logic
        if status_info['has_children']:
            final_status = 'TRANSACTION CONTINUE'
        else:
            final_status = status_info['status']
        
        # Get child bank (Optimized)
        child_bank = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ''
        child_id_str = str(child_trans_id).strip()
        if child_id_str in debited_trans_id_map:
            for c_idx in debited_trans_id_map[child_id_str]:
                if df_main.iloc[c_idx, 5] > current_layer:
                    child_bank = str(df_main.iloc[c_idx, 4])
                    break
        
        all_transactions.append({
            's_no': int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
            'acknowledgement_no': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
            'debited_account': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
            'debited_transaction_id': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
            'bank': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
            'layer': current_layer,
            'credited_account': str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
            'ifsc_code': str(row.iloc[7]) if pd.notna(row.iloc[7]) else '',
            'transaction_date': str(row.iloc[8]) if pd.notna(row.iloc[8]) else '',
            'credited_transaction_id': child_trans_id,
            'transaction_amount': clean_amount(row.iloc[10]),
            'disputed_amount': disputed_amount,
            'reference_no': str(row.iloc[12]) if pd.notna(row.iloc[12]) else '',
            'remarks': str(row.iloc[13]) if pd.notna(row.iloc[13]) else '',
            'action_taken_by': str(row.iloc[14]) if pd.notna(row.iloc[14]) else '',
            'date_of_action': str(row.iloc[15]) if pd.notna(row.iloc[15]) else '',
            'credited_bank': child_bank,
            'status': final_status,
            'updated_amount': status_info['updated_amount'],
            'pending_amount': status_info['pending_amount'],
            'has_children': status_info['has_children'],
            'in_other_sheets': status_info['in_other_sheets']
        })
    
    return jsonify(all_transactions)

@app.route('/api/transaction-details/<trans_id>')
def get_transaction_details(trans_id):
    breakdown = get_transaction_breakdown(trans_id)
    
    details = []
    for item in breakdown:
        details.append({
            'sheet': item['sheet'],
            'amount': item['amount'],
            'bank': '',
            'date': '',
            'remarks': ''
        })
    
    return jsonify(details)

@app.route('/download-report')
def download_report():
    try:
        if df_main is None:
            return jsonify({'success': False, 'message': 'No data loaded'})
        
        all_data = build_hierarchical_data()
        
        if not all_data:
            return jsonify({'success': False, 'message': 'No data to export'})
        
        df = pd.DataFrame(all_data)
        
        # CRITICAL VALIDATION: Check for data loss
        original_count = len(df_main)
        output_count = len(df)
        
        validation_warnings = []
        if output_count != original_count:
            validation_warnings.append(f"⚠️ WARNING: Original file has {original_count} transactions but output has {output_count} transactions!")
            validation_warnings.append(f"⚠️ Difference: {abs(original_count - output_count)} transactions")
        
        # Check if all transaction IDs are present
        original_trans_ids = set(df_main.iloc[:, 9].astype(str).str.strip())
        output_trans_ids = set(df['Credited Transaction ID'].astype(str).str.strip())
        missing_ids = original_trans_ids - output_trans_ids
        
        if missing_ids:
            validation_warnings.append(f"⚠️ WARNING: {len(missing_ids)} transaction IDs missing from output!")
            validation_warnings.append(f"⚠️ Missing IDs: {', '.join(list(missing_ids)[:10])}...")
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Hierarchical Report', index=False)
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = writer.sheets['Hierarchical Report']
            
            # Enable filtering on all columns
            worksheet.auto_filter.ref = worksheet.dimensions
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            column_widths = {
                'A': 8,   # Level
                'B': 8,   # Layer
                'C': 20,  # Debited Trans ID (Filter)
                'D': 50,  # Hierarchy Path
                'E': 8,   # S.No
                'F': 18,  # Debited Account
                'G': 20,  # Debited Transaction ID
                'H': 20,  # Debited Bank
                'I': 18,  # Acknowledgement No
                'J': 18,  # Credited Account
                'K': 20,  # Credited Transaction ID
                'L': 20,  # Credited Bank
                'M': 15,  # IFSC Code
                'N': 20,  # Transaction Date
                'O': 15,  # Transaction Amount
                'P': 15,  # Disputed Amount
                'Q': 18,  # Layerwise Amount Difference
                'R': 15,  # Updated Amount
                'S': 15,  # Pending Amount
                'T': 12,  # Status
                'U': 12,  # Has Children
                'V': 18,  # Found in Other Sheets
                'W': 40,  # Breakdown by Sheet
                'X': 15,  # Reference No
                'Y': 30,  # Remarks
                'Z': 20,  # Action Taken By
                'AA': 20  # Date of Action
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            level_colors = {
                0: 'E3F2FD', 1: 'FFF3E0', 2: 'F3E5F5', 3: 'E8F5E9',
                4: 'FFF9C4', 5: 'FCE4EC', 6: 'E0F2F1'
            }
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                level = df.iloc[row_idx-2]['Level']
                fill_color = level_colors.get(level, 'FFFFFF')
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                # Get status for this row
                status = df.iloc[row_idx-2]['Status']
                
                for cell in row:
                    cell.border = thin_border
                    cell.fill = row_fill
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Color code Status column (column T = 20)
                    if cell.column == 20:  # Status column
                        if status == 'PENDING':
                            cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')  # Light Red
                            cell.font = Font(color='C62828', bold=True)  # Dark Red text
                        elif status == 'COMPLETE':
                            cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # Light Green
                            cell.font = Font(color='2E7D32', bold=True)  # Dark Green text
                        elif status == 'PARTIAL':
                            cell.fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')  # Light Yellow
                            cell.font = Font(color='F57C00', bold=True)  # Dark Orange text
                        elif status == 'TRANSACTION CONTINUE':
                            cell.fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')  # Light Blue
                            cell.font = Font(color='1565C0', bold=True)  # Dark Blue text
                    
                    if cell.column in [15, 16, 17, 18, 19]:  # Amount columns (O, P, Q, R, S)
                        cell.number_format = '₹#,##0.00'
            
            # Summary - Sheet-wise disputed amounts with dynamic formulas
            summary_data = []
            
            # Money Transfer sheet summary (will use formula)
            money_transfer_total = df['Disputed Amount'].sum()
            summary_data.append({
                'Sheet Name': 'Money Transfer to',
                'Total Disputed Amount': money_transfer_total
            })
            
            # Other sheets summary
            for sheet_name, sheet_info in df_other_sheets.items():
                df_sheet = sheet_info['data']
                amount_col = sheet_info['amount_col']
                
                # Calculate total amount without removing duplicates
                if len(df_sheet.columns) > amount_col:
                    total_amount = df_sheet.iloc[:, amount_col].apply(clean_amount).sum()
                else:
                    total_amount = 0
                
                if total_amount > 0:
                    summary_data.append({
                        'Sheet Name': sheet_name,
                        'Total Disputed Amount': total_amount
                    })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False, startrow=0)
            
            summary_ws = writer.sheets['Summary']
            
            # Money Transfer total formula - SUM of all disputed amounts
            summary_ws['B2'] = f"=SUM('Hierarchical Report'!P:P)"
            
            # Header formatting
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            summary_ws.column_dimensions['A'].width = 40
            summary_ws.column_dimensions['B'].width = 25
            
            # Format amounts as currency
            for row_idx in range(2, len(summary_data) + 2):
                cell = summary_ws.cell(row=row_idx, column=2)
                cell.number_format = '₹#,##0.00'
            
            # Grand total with formula
            grand_total_row = len(summary_data) + 2
            summary_ws.cell(row=grand_total_row, column=1, value='GRAND TOTAL')
            summary_ws.cell(row=grand_total_row, column=1).font = Font(bold=True)
            summary_ws.cell(row=grand_total_row, column=2).value = f"=SUM(B2:B{grand_total_row-1})"
            summary_ws.cell(row=grand_total_row, column=2).number_format = '₹#,##0.00'
            summary_ws.cell(row=grand_total_row, column=2).font = Font(bold=True)
            
            # Add charts
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.chart.label import DataLabelList
            
            # Bar Chart
            bar_chart = BarChart()
            bar_chart.title = "Sheet-wise Disputed Amount Distribution"
            bar_chart.style = 10
            bar_chart.y_axis.title = 'Amount (₹)'
            bar_chart.x_axis.title = 'Sheet Name'
            
            data_rows = len(summary_data)
            data = Reference(summary_ws, min_col=2, min_row=1, max_row=data_rows + 1)
            cats = Reference(summary_ws, min_col=1, min_row=2, max_row=data_rows + 1)
            
            bar_chart.add_data(data, titles_from_data=True)
            bar_chart.set_categories(cats)
            bar_chart.height = 15
            bar_chart.width = 25
            
            summary_ws.add_chart(bar_chart, "D2")
            
            # Pie Chart
            pie_chart = PieChart()
            pie_chart.title = "Top Sheets by Disputed Amount"
            pie_chart.style = 10
            
            data = Reference(summary_ws, min_col=2, min_row=2, max_row=min(7, data_rows + 1))
            cats = Reference(summary_ws, min_col=1, min_row=2, max_row=min(7, data_rows + 1))
            
            pie_chart.add_data(data)
            pie_chart.set_categories(cats)
            pie_chart.height = 12
            pie_chart.width = 15
            
            pie_chart.dataLabels = DataLabelList()
            pie_chart.dataLabels.showPercent = True
            
            summary_ws.add_chart(pie_chart, "D32")
            
            # Status Summary with formulas
            status_summary_row = grand_total_row + 3
            summary_ws.cell(row=status_summary_row, column=1, value="Status Summary (Live from Hierarchical Report)")
            summary_ws.cell(row=status_summary_row, column=1).font = Font(bold=True, size=14, color='1F4E78')
            
            status_summary_row += 2
            summary_ws.cell(row=status_summary_row, column=1, value="Status")
            summary_ws.cell(row=status_summary_row, column=2, value="Count")
            summary_ws.cell(row=status_summary_row, column=3, value="Total Amount")
            
            for cell in summary_ws[status_summary_row]:
                cell.fill = header_fill
                cell.font = header_font
            
            status_summary_row += 1
            
            # Status formulas - COUNTIF and SUMIF from Hierarchical Report
            for status in ['PENDING', 'PARTIAL', 'COMPLETE', 'TRANSACTION CONTINUE']:
                summary_ws.cell(row=status_summary_row, column=1, value=status)
                # Count formula
                summary_ws.cell(row=status_summary_row, column=2).value = f"=COUNTIF('Hierarchical Report'!T:T,\"{status}\")"
                # Sum formula
                summary_ws.cell(row=status_summary_row, column=3).value = f"=SUMIF('Hierarchical Report'!T:T,\"{status}\",'Hierarchical Report'!P:P)"
                summary_ws.cell(row=status_summary_row, column=3).number_format = '₹#,##0.00'
                
                # Color code status
                if status == 'PENDING':
                    summary_ws.cell(row=status_summary_row, column=1).fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                elif status == 'COMPLETE':
                    summary_ws.cell(row=status_summary_row, column=1).fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                elif status == 'PARTIAL':
                    summary_ws.cell(row=status_summary_row, column=1).fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')
                elif status == 'TRANSACTION CONTINUE':
                    summary_ws.cell(row=status_summary_row, column=1).fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
                
                status_summary_row += 1
            
            summary_ws.column_dimensions['C'].width = 25
            
            # Bank-wise Summary by Status
            bank_summary_row = status_summary_row + 3
            summary_ws.cell(row=bank_summary_row, column=1, value="Bank-wise Summary by Status")
            summary_ws.cell(row=bank_summary_row, column=1).font = Font(bold=True, size=14, color='1F4E78')
            
            bank_summary_row += 2
            
            # PARTIAL Status - Bank-wise updated Amount
            summary_ws.cell(row=bank_summary_row, column=1, value="PARTIAL Status - updated Amount by Bank")
            summary_ws.cell(row=bank_summary_row, column=1).font = Font(bold=True, size=12)
            summary_ws.cell(row=bank_summary_row, column=1).fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')
            bank_summary_row += 1
            
            summary_ws.cell(row=bank_summary_row, column=1, value="Bank Name")
            summary_ws.cell(row=bank_summary_row, column=2, value="Count")
            summary_ws.cell(row=bank_summary_row, column=3, value="Total updated Amount")
            for cell in summary_ws[bank_summary_row]:
                cell.fill = header_fill
                cell.font = header_font
            bank_summary_row += 1
            
            # Get unique banks for PARTIAL status with formulas
            partial_banks = df[df['Status'] == 'PARTIAL']['Debited Bank'].unique()
            for bank in sorted(partial_banks):
                if bank and str(bank).strip():
                    summary_ws.cell(row=bank_summary_row, column=1, value=str(bank))
                    # Count formula
                    summary_ws.cell(row=bank_summary_row, column=2).value = f"=COUNTIFS('Hierarchical Report'!T:T,\"PARTIAL\",'Hierarchical Report'!H:H,\"{bank}\")"
                    # Sum updated Amount formula
                    summary_ws.cell(row=bank_summary_row, column=3).value = f"=SUMIFS('Hierarchical Report'!R:R,'Hierarchical Report'!T:T,\"PARTIAL\",'Hierarchical Report'!H:H,\"{bank}\")"
                    summary_ws.cell(row=bank_summary_row, column=3).number_format = '₹#,##0.00'
                    bank_summary_row += 1
            
            bank_summary_row += 1
            
            # COMPLETE Status - Bank-wise updated Amount
            summary_ws.cell(row=bank_summary_row, column=1, value="COMPLETE Status - updated Amount by Bank")
            summary_ws.cell(row=bank_summary_row, column=1).font = Font(bold=True, size=12)
            summary_ws.cell(row=bank_summary_row, column=1).fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
            bank_summary_row += 1
            
            summary_ws.cell(row=bank_summary_row, column=1, value="Bank Name")
            summary_ws.cell(row=bank_summary_row, column=2, value="Count")
            summary_ws.cell(row=bank_summary_row, column=3, value="Total updated Amount")
            for cell in summary_ws[bank_summary_row]:
                cell.fill = header_fill
                cell.font = header_font
            bank_summary_row += 1
            
            # Get unique banks for COMPLETE status with formulas
            complete_banks = df[df['Status'] == 'COMPLETE']['Debited Bank'].unique()
            for bank in sorted(complete_banks):
                if bank and str(bank).strip():
                    summary_ws.cell(row=bank_summary_row, column=1, value=str(bank))
                    # Count formula
                    summary_ws.cell(row=bank_summary_row, column=2).value = f"=COUNTIFS('Hierarchical Report'!T:T,\"COMPLETE\",'Hierarchical Report'!H:H,\"{bank}\")"
                    # Sum updated Amount formula
                    summary_ws.cell(row=bank_summary_row, column=3).value = f"=SUMIFS('Hierarchical Report'!R:R,'Hierarchical Report'!T:T,\"COMPLETE\",'Hierarchical Report'!H:H,\"{bank}\")"
                    summary_ws.cell(row=bank_summary_row, column=3).number_format = '₹#,##0.00'
                    bank_summary_row += 1
            
            bank_summary_row += 1
            
            # PENDING Status - Bank-wise Pending Amount
            summary_ws.cell(row=bank_summary_row, column=1, value="PENDING Status - Pending Amount by Bank")
            summary_ws.cell(row=bank_summary_row, column=1).font = Font(bold=True, size=12)
            summary_ws.cell(row=bank_summary_row, column=1).fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
            bank_summary_row += 1
            
            summary_ws.cell(row=bank_summary_row, column=1, value="Bank Name")
            summary_ws.cell(row=bank_summary_row, column=2, value="Count")
            summary_ws.cell(row=bank_summary_row, column=3, value="Total Pending Amount")
            for cell in summary_ws[bank_summary_row]:
                cell.fill = header_fill
                cell.font = header_font
            bank_summary_row += 1
            
            # Get unique banks for PENDING status with formulas
            pending_banks = df[df['Status'] == 'PENDING']['Debited Bank'].unique()
            for bank in sorted(pending_banks):
                if bank and str(bank).strip():
                    summary_ws.cell(row=bank_summary_row, column=1, value=str(bank))
                    # Count formula
                    summary_ws.cell(row=bank_summary_row, column=2).value = f"=COUNTIFS('Hierarchical Report'!T:T,\"PENDING\",'Hierarchical Report'!H:H,\"{bank}\")"
                    # Sum Pending Amount formula
                    summary_ws.cell(row=bank_summary_row, column=3).value = f"=SUMIFS('Hierarchical Report'!S:S,'Hierarchical Report'!T:T,\"PENDING\",'Hierarchical Report'!H:H,\"{bank}\")"
                    summary_ws.cell(row=bank_summary_row, column=3).number_format = '₹#,##0.00'
                    bank_summary_row += 1
            
            summary_ws.column_dimensions['A'].width = 40
            summary_ws.column_dimensions['B'].width = 15
            summary_ws.column_dimensions['C'].width = 25
            
            # Add validation warnings if any
            if validation_warnings:
                warning_row = bank_summary_row + 3
                summary_ws.cell(row=warning_row, column=1, value="⚠️ DATA VALIDATION WARNINGS")
                summary_ws.cell(row=warning_row, column=1).font = Font(bold=True, size=14, color='FF0000')
                warning_row += 2
                
                for warning in validation_warnings:
                    summary_ws.cell(row=warning_row, column=1, value=warning)
                    summary_ws.cell(row=warning_row, column=1).font = Font(color='FF0000', bold=True)
                    summary_ws.cell(row=warning_row, column=1).fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                    warning_row += 1
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Transaction_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/download-leaf-nodes')
def download_leaf_nodes():
    """Download report with only leaf nodes (transactions with no children)"""
    try:
        if df_main is None:
            return jsonify({'success': False, 'message': 'No data loaded'})
        
        # Build full hierarchy
        all_data = build_hierarchical_data()
        
        if not all_data:
            return jsonify({'success': False, 'message': 'No data to export'})
        
        # Filter only leaf nodes (Has Children = 'No')
        df_full = pd.DataFrame(all_data)
        df = df_full[df_full['Has Children'] == 'No'].copy()
        
        if len(df) == 0:
            return jsonify({'success': False, 'message': 'No leaf node transactions found'})
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Leaf Nodes Report', index=False)
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = writer.sheets['Leaf Nodes Report']
            
            # Enable filtering
            worksheet.auto_filter.ref = worksheet.dimensions
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Same column widths as main report
            column_widths = {
                'A': 8, 'B': 8, 'C': 20, 'D': 50, 'E': 8, 'F': 18, 'G': 18, 'H': 20,
                'I': 20, 'J': 18, 'K': 20, 'L': 20, 'M': 15, 'N': 20, 'O': 15,
                'P': 15, 'Q': 18, 'R': 15, 'S': 15, 'T': 12, 'U': 12, 'V': 18,
                'W': 40, 'X': 15, 'Y': 30, 'Z': 20, 'AA': 20
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                status = df.iloc[row_idx-2]['Status']
                
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Color code Status column
                    if cell.column == 20:  # Status column (now column T = 20)
                        if status == 'PENDING':
                            cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
                            cell.font = Font(color='C62828', bold=True)
                        elif status == 'COMPLETE':
                            cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
                            cell.font = Font(color='2E7D32', bold=True)
                        elif status == 'PARTIAL':
                            cell.fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')
                            cell.font = Font(color='F57C00', bold=True)
                        elif status == 'TRANSACTION CONTINUE':
                            cell.fill = PatternFill(start_color='BBDEFB', end_color='BBDEFB', fill_type='solid')
                            cell.font = Font(color='1565C0', bold=True)
                    
                    if cell.column in [15, 16, 17, 18, 19]:  # Amount columns
                        cell.number_format = '₹#,##0.00'
            
            # Summary for leaf nodes
            summary_data = []
            
            # Count by status
            for status in ['PENDING', 'PARTIAL', 'COMPLETE']:
                status_data = df[df['Status'] == status]
                if len(status_data) > 0:
                    summary_data.append({
                        'Category': f'{status} Transactions',
                        'Count': len(status_data),
                        'Total Disputed': status_data['Disputed Amount'].sum(),
                        'Total Pending': status_data['Pending Amount'].sum()
                    })
            
            summary_data.append({})
            summary_data.append({
                'Category': 'TOTAL LEAF NODES',
                'Count': len(df),
                'Total Disputed': df['Disputed Amount'].sum(),
                'Total Pending': df['Pending Amount'].sum()
            })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            summary_ws = writer.sheets['Summary']
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 15
            summary_ws.column_dimensions['C'].width = 20
            summary_ws.column_dimensions['D'].width = 20
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Leaf_Nodes_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/download-partial')
def download_partial():
    """Download report with only PARTIAL status transactions"""
    try:
        if df_main is None:
            return jsonify({'success': False, 'message': 'No data loaded'})
        
        # Build full hierarchy
        all_data = build_hierarchical_data()
        
        if not all_data:
            return jsonify({'success': False, 'message': 'No data to export'})
        
        # Filter only PARTIAL status
        df_full = pd.DataFrame(all_data)
        df = df_full[df_full['Status'] == 'PARTIAL'].copy()
        
        if len(df) == 0:
            return jsonify({'success': False, 'message': 'No PARTIAL status transactions found'})
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Partial Status Report', index=False)
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = writer.sheets['Partial Status Report']
            
            # Enable filtering
            worksheet.auto_filter.ref = worksheet.dimensions
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Same column widths as main report
            column_widths = {
                'A': 8, 'B': 8, 'C': 20, 'D': 50, 'E': 8, 'F': 18, 'G': 18, 'H': 20,
                'I': 20, 'J': 18, 'K': 20, 'L': 20, 'M': 15, 'N': 20, 'O': 15,
                'P': 15, 'Q': 18, 'R': 15, 'S': 15, 'T': 12, 'U': 12, 'V': 18,
                'W': 40, 'X': 15, 'Y': 30, 'Z': 20, 'AA': 20
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                status = df.iloc[row_idx-2]['Status']
                
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Color code Status column (PARTIAL = Yellow)
                    if cell.column == 20:  # Status column
                        cell.fill = PatternFill(start_color='FFE082', end_color='FFE082', fill_type='solid')
                        cell.font = Font(color='F57C00', bold=True)
                    
                    if cell.column in [15, 16, 17, 18, 19]:  # Amount columns
                        cell.number_format = '₹#,##0.00'
            
            # Summary for partial transactions
            summary_data = []
            summary_data.append({
                'Category': 'PARTIAL Transactions',
                'Count': len(df),
                'Total Disputed': df['Disputed Amount'].sum(),
                'Total updated': df['Updated Amount'].sum(),
                'Total Pending': df['Pending Amount'].sum()
            })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            summary_ws = writer.sheets['Summary']
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 15
            summary_ws.column_dimensions['C'].width = 20
            summary_ws.column_dimensions['D'].width = 20
            summary_ws.column_dimensions['E'].width = 20
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Partial_Status_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

if __name__ == '__main__':
    app.run(debug=False, port=5000, host='127.0.0.1')
