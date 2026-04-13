from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import sqlite3
import os
from datetime import datetime
from io import BytesIO

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('reports', exist_ok=True)

def clean_amount(value):
    """Convert amount string to float"""
    if pd.isna(value) or value == '':
        return 0.0
    if isinstance(value, str):
        value = value.replace(',', '').replace('₹', '').strip()
    try:
        return float(value)
    except:
        return 0.0

def init_database():
    """Initialize SQLite database"""
    conn = sqlite3.connect('transactions.db')
    cursor = conn.cursor()
    
    # Main transactions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            s_no INTEGER,
            acknowledgement_no TEXT,
            parent_account TEXT,
            parent_transaction_id TEXT,
            bank TEXT,
            layer INTEGER,
            child_account TEXT,
            ifsc_code TEXT,
            transaction_date TEXT,
            child_transaction_id TEXT,
            transaction_amount REAL,
            disputed_amount REAL,
            reference_no TEXT,
            remarks TEXT,
            action_taken_by TEXT,
            date_of_action TEXT,
            pisnodal INTEGER
        )
    ''')
    
    # Other sheets table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS other_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sheet_name TEXT,
            transaction_id TEXT,
            account_no TEXT,
            amount REAL,
            bank TEXT,
            transaction_date TEXT,
            remarks TEXT,
            raw_data TEXT
        )
    ''')
    
    conn.commit()
    conn.close()

def clear_database():
    """Clear all data from database"""
    conn = sqlite3.connect('transactions.db')
    cursor = conn.cursor()
    cursor.execute('DELETE FROM transactions')
    cursor.execute('DELETE FROM other_transactions')
    conn.commit()
    conn.close()

def process_excel_file(filepath):
    """Process Excel file and load into database"""
    clear_database()
    conn = sqlite3.connect('transactions.db')
    
    xl = pd.ExcelFile(filepath)
    
    # Process Money Transfer sheet
    money_transfer_sheet = None
    for sheet in xl.sheet_names:
        if 'money transfer' in sheet.lower():
            money_transfer_sheet = sheet
            break
    
    if not money_transfer_sheet:
        conn.close()
        return False, "Money Transfer sheet not found"
    
    df_main = pd.read_excel(filepath, sheet_name=money_transfer_sheet)
    
    # Insert main transactions
    for _, row in df_main.iterrows():
        conn.execute('''
            INSERT INTO transactions (
                s_no, acknowledgement_no, parent_account, parent_transaction_id,
                bank, layer, child_account, ifsc_code, transaction_date,
                child_transaction_id, transaction_amount, disputed_amount,
                reference_no, remarks, action_taken_by, date_of_action, pisnodal
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            int(row.iloc[0]) if pd.notna(row.iloc[0]) else None,
            str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
            str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
            str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
            str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
            int(row.iloc[5]) if pd.notna(row.iloc[5]) else 1,
            str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
            str(row.iloc[7]) if pd.notna(row.iloc[7]) else '',
            str(row.iloc[8]) if pd.notna(row.iloc[8]) else '',
            str(row.iloc[9]) if pd.notna(row.iloc[9]) else '',
            clean_amount(row.iloc[10]),
            clean_amount(row.iloc[11]),
            str(row.iloc[12]) if pd.notna(row.iloc[12]) else '',
            str(row.iloc[13]) if pd.notna(row.iloc[13]) else '',
            str(row.iloc[14]) if pd.notna(row.iloc[14]) else '',
            str(row.iloc[15]) if pd.notna(row.iloc[15]) else '',
            int(row.iloc[16]) if pd.notna(row.iloc[16]) else 0
        ))
    
    # Process other sheets
    for sheet in xl.sheet_names:
        if sheet == money_transfer_sheet:
            continue
        
        df_other = pd.read_excel(filepath, sheet_name=sheet)
        
        # Determine amount column (F=5 or I=8 for Cheque)
        amount_col_idx = 8 if 'cheque' in sheet.lower() else 5
        
        for _, row in df_other.iterrows():
            # Transaction ID is usually in column 3 (index 3)
            trans_id = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ''
            account_no = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ''
            amount = clean_amount(row.iloc[amount_col_idx]) if len(row) > amount_col_idx else 0.0
            
            # Try to get bank and date
            bank = ''
            trans_date = ''
            remarks = ''
            
            for i, val in enumerate(row):
                val_str = str(val).lower()
                if 'bank' in val_str and bank == '':
                    bank = str(val)
                if '/' in str(val) and trans_date == '':
                    trans_date = str(val)
            
            conn.execute('''
                INSERT INTO other_transactions (
                    sheet_name, transaction_id, account_no, amount,
                    bank, transaction_date, remarks, raw_data
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                sheet,
                trans_id,
                account_no,
                amount,
                bank,
                trans_date,
                remarks,
                str(row.to_dict())
            ))
    
    conn.commit()
    conn.close()
    return True, "File processed successfully"

def calculate_transaction_status(child_trans_id, disputed_amount, current_layer=None):
    """Calculate status and pending amount for a transaction"""
    conn = sqlite3.connect('transactions.db')
    cursor = conn.cursor()
    
    # Check if transaction has children in Money Transfer sheet
    # Exclude same layer to avoid self-reference
    if current_layer:
        cursor.execute('''
            SELECT COUNT(*) FROM transactions 
            WHERE parent_transaction_id = ? AND layer > ?
        ''', (child_trans_id, current_layer))
    else:
        cursor.execute('''
            SELECT COUNT(*) FROM transactions 
            WHERE parent_transaction_id = ? AND child_transaction_id != parent_transaction_id
        ''', (child_trans_id,))
    has_children = cursor.fetchone()[0] > 0
    
    # Get sum of amounts from other sheets
    cursor.execute('''
        SELECT COALESCE(SUM(amount), 0) FROM other_transactions 
        WHERE transaction_id = ?
    ''', (child_trans_id,))
    other_sheets_total = cursor.fetchone()[0]
    
    # Check if transaction exists in other sheets
    cursor.execute('''
        SELECT COUNT(*) FROM other_transactions 
        WHERE transaction_id = ?
    ''', (child_trans_id,))
    in_other_sheets = cursor.fetchone()[0] > 0
    
    conn.close()
    
    pending_amount = disputed_amount - other_sheets_total
    
    # Determine status
    if not has_children and not in_other_sheets:
        status = 'PENDING'
    elif pending_amount > 0:
        status = 'PARTIAL'
    else:
        status = 'COMPLETE'
    
    return {
        'status': status,
        'updated_amount': other_sheets_total,
        'pending_amount': max(0, pending_amount),
        'has_children': has_children,
        'in_other_sheets': in_other_sheets
    }

def get_layer_transactions(layer=1, parent_trans_id=None):
    """Get transactions for a specific layer"""
    conn = sqlite3.connect('transactions.db')
    cursor = conn.cursor()
    
    if layer == 1:
        query = '''
            SELECT * FROM transactions 
            WHERE layer = 1
            ORDER BY s_no
        '''
        cursor.execute(query)
    else:
        query = '''
            SELECT * FROM transactions 
            WHERE parent_transaction_id = ?
            ORDER BY s_no
        '''
        cursor.execute(query, (parent_trans_id,))
    
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    
    transactions = []
    for row in rows:
        trans_dict = dict(zip(columns, row))
        
        # Calculate status
        status_info = calculate_transaction_status(
            trans_dict['child_transaction_id'],
            trans_dict['disputed_amount']
        )
        
        trans_dict.update(status_info)
        
        # Get child bank
        cursor.execute('''
            SELECT bank FROM transactions 
            WHERE parent_transaction_id = ? 
            LIMIT 1
        ''', (trans_dict['child_transaction_id'],))
        child_bank_result = cursor.fetchone()
        trans_dict['child_bank'] = child_bank_result[0] if child_bank_result else trans_dict['bank']
        
        transactions.append(trans_dict)
    
    conn.close()
    return transactions

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Only Excel files allowed'})
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.xlsx')
    file.save(filepath)
    
    success, message = process_excel_file(filepath)
    return jsonify({'success': success, 'message': message})

@app.route('/api/transactions')
def get_transactions():
    layer = int(request.args.get('layer', 1))
    parent_id = request.args.get('parent_id', None)
    
    transactions = get_layer_transactions(layer, parent_id)
    return jsonify(transactions)

@app.route('/api/transaction-details/<trans_id>')
def get_transaction_details(trans_id):
    """Get detailed breakdown of a transaction from all sheets"""
    conn = sqlite3.connect('transactions.db')
    cursor = conn.cursor()
    
    # Get from other sheets
    cursor.execute('''
        SELECT sheet_name, amount, bank, transaction_date, remarks 
        FROM other_transactions 
        WHERE transaction_id = ?
    ''', (trans_id,))
    
    details = []
    for row in cursor.fetchall():
        details.append({
            'sheet': row[0],
            'amount': row[1],
            'bank': row[2],
            'date': row[3],
            'remarks': row[4]
        })
    
    conn.close()
    return jsonify(details)

def build_hierarchical_data(parent_id=None, layer=1, parent_path="", level=0, visited=None):
    """Recursively build hierarchical transaction data"""
    if visited is None:
        visited = set()
    
    # Prevent infinite recursion
    if level > 10:  # Max 10 levels deep
        return []
    
    conn = sqlite3.connect('transactions.db')
    cursor = conn.cursor()
    
    if layer == 1:
        query = 'SELECT * FROM transactions WHERE layer = 1 ORDER BY s_no'
        cursor.execute(query)
    else:
        # Prevent circular references
        if parent_id in visited:
            conn.close()
            return []
        visited.add(parent_id)
        
        # Get children - must be in higher layer to avoid self-reference
        query = '''SELECT * FROM transactions 
                   WHERE parent_transaction_id = ? 
                   AND layer > (SELECT layer FROM transactions WHERE child_transaction_id = ? LIMIT 1)
                   ORDER BY s_no'''
        cursor.execute(query, (parent_id, parent_id))
    
    columns = [desc[0] for desc in cursor.description]
    rows = cursor.fetchall()
    
    all_data = []
    
    for row in rows:
        trans_dict = dict(zip(columns, row))
        
        # Calculate status
        status_info = calculate_transaction_status(
            trans_dict['child_transaction_id'],
            trans_dict['disputed_amount'],
            trans_dict['layer']
        )
        
        # Get child bank
        cursor.execute('''
            SELECT bank FROM transactions 
            WHERE parent_transaction_id = ? AND layer > ?
            LIMIT 1
        ''', (trans_dict['child_transaction_id'], trans_dict['layer']))
        child_bank_result = cursor.fetchone()
        child_bank = child_bank_result[0] if child_bank_result else trans_dict['bank']
        
        # Get breakdown from other sheets
        cursor.execute('''
            SELECT sheet_name, amount 
            FROM other_transactions 
            WHERE transaction_id = ?
        ''', (trans_dict['child_transaction_id'],))
        breakdown = cursor.fetchall()
        breakdown_text = "; ".join([f"{sheet}: ₹{amt:,.2f}" for sheet, amt in breakdown]) if breakdown else "None"
        
        # Build current path
        current_path = f"{parent_path} → {trans_dict['child_transaction_id']}" if parent_path else trans_dict['child_transaction_id']
        
        # Create row data
        row_data = {
            'Level': level,
            'Hierarchy Path': current_path,
            'Layer': trans_dict['layer'],
            'S.No': trans_dict['s_no'],
            'Acknowledgement No': trans_dict['acknowledgement_no'],
            'Parent Account': trans_dict['parent_account'],
            'Parent Transaction ID': trans_dict['parent_transaction_id'],
            'Parent Bank': trans_dict['bank'],
            'Child Account': trans_dict['child_account'],
            'Child Transaction ID': trans_dict['child_transaction_id'],
            'Child Bank': child_bank,
            'IFSC Code': trans_dict['ifsc_code'],
            'Transaction Date': trans_dict['transaction_date'],
            'Transaction Amount': trans_dict['transaction_amount'],
            'Disputed Amount': trans_dict['disputed_amount'],
            'Updated Amount': status_info['updated_amount'],
            'Pending Amount': status_info['pending_amount'],
            'Status': status_info['status'],
            'Has Children': 'Yes' if status_info['has_children'] else 'No',
            'Found in Other Sheets': 'Yes' if status_info['in_other_sheets'] else 'No',
            'Breakdown by Sheet': breakdown_text,
            'Reference No': trans_dict['reference_no'],
            'Remarks': trans_dict['remarks'],
            'Action Taken By': trans_dict['action_taken_by'],
            'Date of Action': trans_dict['date_of_action']
        }
        
        all_data.append(row_data)
        
        # Recursively get children
        if status_info['has_children']:
            children_data = build_hierarchical_data(
                trans_dict['child_transaction_id'],
                trans_dict['layer'] + 1,
                current_path,
                level + 1,
                visited.copy()  # Pass a copy to avoid affecting siblings
            )
            all_data.extend(children_data)
    
    conn.close()
    return all_data

@app.route('/download-report')
def download_report():
    """Generate and download hierarchical Excel report"""
    try:
        # Build hierarchical data
        all_data = build_hierarchical_data()
        
        if not all_data:
            return jsonify({'success': False, 'message': 'No data to export'})
        
        # Create DataFrame
        df = pd.DataFrame(all_data)
        
        # Create Excel file in memory
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Main hierarchical sheet
            df.to_excel(writer, sheet_name='Hierarchical Report', index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Hierarchical Report']
            
            # Format headers
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set column widths
            column_widths = {
                'A': 8,   # Level
                'B': 50,  # Hierarchy Path
                'C': 8,   # Layer
                'D': 8,   # S.No
                'E': 18,  # Acknowledgement No
                'F': 18,  # Parent Account
                'G': 20,  # Parent Transaction ID
                'H': 20,  # Parent Bank
                'I': 18,  # Child Account
                'J': 20,  # Child Transaction ID
                'K': 20,  # Child Bank
                'L': 15,  # IFSC Code
                'M': 20,  # Transaction Date
                'N': 15,  # Transaction Amount
                'O': 15,  # Disputed Amount
                'P': 15,  # Updated Amount
                'Q': 15,  # Pending Amount
                'R': 12,  # Status
                'S': 12,  # Has Children
                'T': 18,  # Found in Other Sheets
                'U': 40,  # Breakdown by Sheet
                'V': 15,  # Reference No
                'W': 30,  # Remarks
                'X': 20,  # Action Taken By
                'Y': 20   # Date of Action
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Format data rows
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            # Color code by level
            level_colors = {
                0: 'E3F2FD',  # Light blue
                1: 'FFF3E0',  # Light orange
                2: 'F3E5F5',  # Light purple
                3: 'E8F5E9',  # Light green
                4: 'FFF9C4',  # Light yellow
                5: 'FCE4EC',  # Light pink
                6: 'E0F2F1',  # Light teal
            }
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                level = df.iloc[row_idx-2]['Level']
                fill_color = level_colors.get(level, 'FFFFFF')
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                for cell in row:
                    cell.border = thin_border
                    cell.fill = row_fill
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Format amounts
                    if cell.column in [14, 15, 16, 17]:  # Amount columns
                        cell.number_format = '₹#,##0.00'
            
            # Add summary sheet
            conn = sqlite3.connect('transactions.db')
            
            # Summary statistics
            summary_data = []
            
            # Layer-wise summary
            for layer in range(1, 8):
                cursor = conn.cursor()
                cursor.execute('''
                    SELECT 
                        COUNT(*) as count,
                        SUM(disputed_amount) as total_disputed,
                        SUM(transaction_amount) as total_transaction
                    FROM transactions 
                    WHERE layer = ?
                ''', (layer,))
                result = cursor.fetchone()
                if result[0] > 0:
                    summary_data.append({
                        'Layer': f'Layer {layer}',
                        'Transaction Count': result[0],
                        'Total Disputed Amount': result[1] or 0,
                        'Total Transaction Amount': result[2] or 0
                    })
            
            # Status summary
            cursor.execute('''
                SELECT 
                    COUNT(*) as total_transactions,
                    SUM(disputed_amount) as total_disputed
                FROM transactions
            ''')
            total_trans, total_disputed = cursor.fetchone()
            
            summary_data.append({})  # Empty row
            summary_data.append({
                'Layer': 'TOTAL',
                'Transaction Count': total_trans,
                'Total Disputed Amount': total_disputed or 0,
                'Total Transaction Amount': ''
            })
            
            conn.close()
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Format summary sheet
            summary_ws = writer.sheets['Summary']
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 20
            summary_ws.column_dimensions['C'].width = 25
            summary_ws.column_dimensions['D'].width = 25
        
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Transaction_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error generating report: {str(e)}'})

if __name__ == '__main__':
    init_database()
    app.run(debug=True, port=5000)
