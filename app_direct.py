from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
import os
from datetime import datetime
from io import BytesIO

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Global variables to store data
df_main = None
df_other_sheets = {}
current_file = None

def clean_amount(value):
    """Convert amount string to float"""
    if pd.isna(value) or value == '':
        return 0.0
    if isinstance(value, str):
        value = value.replace(',', '').replace('₹', '').strip()
    try:
        return float(value)
    except:
        return 0.0

def process_excel_file(filepath):
    """Process Excel file and load into memory"""
    global df_main, df_other_sheets, current_file
    
    try:
        xl = pd.ExcelFile(filepath)
        
        # Find Money Transfer sheet
        money_transfer_sheet = None
        for sheet in xl.sheet_names:
            if 'money transfer' in sheet.lower():
                money_transfer_sheet = sheet
                break
        
        if not money_transfer_sheet:
            return False, "Money Transfer sheet not found"
        
        # Load main sheet
        df_main = pd.read_excel(filepath, sheet_name=money_transfer_sheet)
        
        # Clean column names
        df_main.columns = df_main.columns.str.strip()
        
        # Load other sheets
        df_other_sheets = {}
        for sheet in xl.sheet_names:
            if sheet == money_transfer_sheet:
                continue
            
            df_sheet = pd.read_excel(filepath, sheet_name=sheet)
            df_sheet.columns = df_sheet.columns.str.strip()
            
            # Determine amount column
            amount_col_idx = 8 if 'cheque' in sheet.lower() else 5
            
            # Store sheet data
            df_other_sheets[sheet] = {
                'data': df_sheet,
                'amount_col': amount_col_idx
            }
        
        current_file = os.path.basename(filepath)
        return True, f"File processed successfully. Found {len(df_other_sheets)} additional sheets."
        
    except Exception as e:
        return False, f"Error processing file: {str(e)}"

def get_transaction_breakdown(trans_id):
    """Get breakdown of a transaction from all sheets"""
    breakdown = []
    
    for sheet_name, sheet_info in df_other_sheets.items():
        df = sheet_info['data']
        amount_col = sheet_info['amount_col']
        
        # Transaction ID is usually in column 3
        if len(df.columns) > 3:
            matches = df[df.iloc[:, 3].astype(str) == str(trans_id)]
            
            for _, row in matches.iterrows():
                if len(row) > amount_col:
                    amount = clean_amount(row.iloc[amount_col])
                    breakdown.append({
                        'sheet': sheet_name,
                        'amount': amount
                    })
    
    return breakdown

def calculate_transaction_status(child_trans_id, disputed_amount, current_layer):
    """Calculate status for a transaction"""
    # Check if has children
    has_children = False
    if df_main is not None:
        children = df_main[
            (df_main.iloc[:, 3].astype(str) == str(child_trans_id)) & 
            (df_main.iloc[:, 5] > current_layer)
        ]
        has_children = len(children) > 0
    
    # Get breakdown from other sheets
    breakdown = get_transaction_breakdown(child_trans_id)
    updated_amount = sum(item['amount'] for item in breakdown)
    in_other_sheets = len(breakdown) > 0
    
    pending_amount = disputed_amount - updated_amount
    
    # Determine status
    if not has_children and not in_other_sheets:
        status = 'PENDING'
    elif pending_amount > 0:
        status = 'PARTIAL'
    else:
        status = 'COMPLETE'
    
    return {
        'status': status,
        'updated_amount': updated_amount,
        'pending_amount': max(0, pending_amount),
        'has_children': has_children,
        'in_other_sheets': in_other_sheets
    }

def get_layer_transactions(layer=1, parent_trans_id=None):
    """Get transactions for a specific layer"""
    if df_main is None:
        return []
    
    if layer == 1:
        filtered = df_main[df_main.iloc[:, 5] == 1]
    else:
        # Get children of parent
        parent_layer = df_main[df_main.iloc[:, 9].astype(str) == str(parent_trans_id)].iloc[:, 5].iloc[0] if len(df_main[df_main.iloc[:, 9].astype(str) == str(parent_trans_id)]) > 0 else 0
        filtered = df_main[
            (df_main.iloc[:, 3].astype(str) == str(parent_trans_id)) & 
            (df_main.iloc[:, 5] > parent_layer)
        ]
    
    transactions = []
    for _, row in filtered.iterrows():
        child_trans_id = str(row.iloc[9]) if pd.notna(row.iloc[9]) else ''
        disputed_amount = clean_amount(row.iloc[11])
        current_layer = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 1
        
        status_info = calculate_transaction_status(child_trans_id, disputed_amount, current_layer)
        
        # Get child bank
        child_bank = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ''
        children = df_main[
            (df_main.iloc[:, 3].astype(str) == child_trans_id) & 
            (df_main.iloc[:, 5] > current_layer)
        ]
        if len(children) > 0:
            child_bank = str(children.iloc[0, 4])
        
        transactions.append({
            's_no': int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
            'acknowledgement_no': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
            'parent_account': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
            'parent_transaction_id': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
            'bank': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
            'layer': current_layer,
            'child_account': str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
            'ifsc_code': str(row.iloc[7]) if pd.notna(row.iloc[7]) else '',
            'transaction_date': str(row.iloc[8]) if pd.notna(row.iloc[8]) else '',
            'child_transaction_id': child_trans_id,
            'transaction_amount': clean_amount(row.iloc[10]),
            'disputed_amount': disputed_amount,
            'reference_no': str(row.iloc[12]) if pd.notna(row.iloc[12]) else '',
            'remarks': str(row.iloc[13]) if pd.notna(row.iloc[13]) else '',
            'action_taken_by': str(row.iloc[14]) if pd.notna(row.iloc[14]) else '',
            'date_of_action': str(row.iloc[15]) if pd.notna(row.iloc[15]) else '',
            'child_bank': child_bank,
            **status_info
        })
    
    return transactions

def build_hierarchical_data(parent_id=None, layer=1, parent_path="", level=0, visited=None):
    """Build hierarchical data for Excel export"""
    if visited is None:
        visited = set()
    
    if level > 10 or df_main is None:
        return []
    
    if layer == 1:
        filtered = df_main[df_main.iloc[:, 5] == 1]
    else:
        if parent_id in visited:
            return []
        visited.add(parent_id)
        
        parent_layer = df_main[df_main.iloc[:, 9].astype(str) == str(parent_id)].iloc[:, 5].iloc[0] if len(df_main[df_main.iloc[:, 9].astype(str) == str(parent_id)]) > 0 else 0
        filtered = df_main[
            (df_main.iloc[:, 3].astype(str) == str(parent_id)) & 
            (df_main.iloc[:, 5] > parent_layer)
        ]
    
    all_data = []
    
    for _, row in filtered.iterrows():
        child_trans_id = str(row.iloc[9]) if pd.notna(row.iloc[9]) else ''
        disputed_amount = clean_amount(row.iloc[11])
        current_layer = int(row.iloc[5]) if pd.notna(row.iloc[5]) else 1
        
        status_info = calculate_transaction_status(child_trans_id, disputed_amount, current_layer)
        
        # Get breakdown
        breakdown = get_transaction_breakdown(child_trans_id)
        breakdown_text = "; ".join([f"{item['sheet']}: ₹{item['amount']:,.2f}" for item in breakdown]) if breakdown else "None"
        
        # Get child bank
        child_bank = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ''
        children = df_main[
            (df_main.iloc[:, 3].astype(str) == child_trans_id) & 
            (df_main.iloc[:, 5] > current_layer)
        ]
        if len(children) > 0:
            child_bank = str(children.iloc[0, 4])
        
        current_path = f"{parent_path} → {child_trans_id}" if parent_path else child_trans_id
        
        row_data = {
            'Level': level,
            'Hierarchy Path': current_path,
            'Layer': current_layer,
            'S.No': int(row.iloc[0]) if pd.notna(row.iloc[0]) else 0,
            'Acknowledgement No': str(row.iloc[1]) if pd.notna(row.iloc[1]) else '',
            'Parent Account': str(row.iloc[2]) if pd.notna(row.iloc[2]) else '',
            'Parent Transaction ID': str(row.iloc[3]) if pd.notna(row.iloc[3]) else '',
            'Parent Bank': str(row.iloc[4]) if pd.notna(row.iloc[4]) else '',
            'Child Account': str(row.iloc[6]) if pd.notna(row.iloc[6]) else '',
            'Child Transaction ID': child_trans_id,
            'Child Bank': child_bank,
            'IFSC Code': str(row.iloc[7]) if pd.notna(row.iloc[7]) else '',
            'Transaction Date': str(row.iloc[8]) if pd.notna(row.iloc[8]) else '',
            'Transaction Amount': clean_amount(row.iloc[10]),
            'Disputed Amount': disputed_amount,
            'Updated Amount': status_info['updated_amount'],
            'Pending Amount': status_info['pending_amount'],
            'Status': status_info['status'],
            'Has Children': 'Yes' if status_info['has_children'] else 'No',
            'Found in Other Sheets': 'Yes' if status_info['in_other_sheets'] else 'No',
            'Breakdown by Sheet': breakdown_text,
            'Reference No': str(row.iloc[12]) if pd.notna(row.iloc[12]) else '',
            'Remarks': str(row.iloc[13]) if pd.notna(row.iloc[13]) else '',
            'Action Taken By': str(row.iloc[14]) if pd.notna(row.iloc[14]) else '',
            'Date of Action': str(row.iloc[15]) if pd.notna(row.iloc[15]) else ''
        }
        
        all_data.append(row_data)
        
        if status_info['has_children']:
            children_data = build_hierarchical_data(
                child_trans_id,
                current_layer + 1,
                current_path,
                level + 1,
                visited.copy()
            )
            all_data.extend(children_data)
    
    return all_data

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': 'No file selected'})
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'message': 'Only Excel files allowed'})
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'current.xlsx')
    file.save(filepath)
    
    success, message = process_excel_file(filepath)
    return jsonify({'success': success, 'message': message})

@app.route('/api/transactions')
def get_transactions():
    layer = int(request.args.get('layer', 1))
    parent_id = request.args.get('parent_id', None)
    
    transactions = get_layer_transactions(layer, parent_id)
    return jsonify(transactions)

@app.route('/api/transaction-details/<trans_id>')
def get_transaction_details(trans_id):
    """Get detailed breakdown of a transaction"""
    breakdown = get_transaction_breakdown(trans_id)
    
    details = []
    for item in breakdown:
        details.append({
            'sheet': item['sheet'],
            'amount': item['amount'],
            'bank': '',
            'date': '',
            'remarks': ''
        })
    
    return jsonify(details)

@app.route('/download-report')
def download_report():
    """Generate and download hierarchical Excel report"""
    try:
        if df_main is None:
            return jsonify({'success': False, 'message': 'No data loaded'})
        
        all_data = build_hierarchical_data()
        
        if not all_data:
            return jsonify({'success': False, 'message': 'No data to export'})
        
        df = pd.DataFrame(all_data)
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Hierarchical Report', index=False)
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            workbook = writer.book
            worksheet = writer.sheets['Hierarchical Report']
            
            # Format headers
            header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True, size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Set column widths
            column_widths = {
                'A': 8, 'B': 50, 'C': 8, 'D': 8, 'E': 18, 'F': 18, 'G': 20, 'H': 20,
                'I': 18, 'J': 20, 'K': 20, 'L': 15, 'M': 20, 'N': 15, 'O': 15,
                'P': 15, 'Q': 15, 'R': 12, 'S': 12, 'T': 18, 'U': 40, 'V': 15,
                'W': 30, 'X': 20, 'Y': 20
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Format data rows
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            level_colors = {
                0: 'E3F2FD', 1: 'FFF3E0', 2: 'F3E5F5', 3: 'E8F5E9',
                4: 'FFF9C4', 5: 'FCE4EC', 6: 'E0F2F1'
            }
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), start=2):
                level = df.iloc[row_idx-2]['Level']
                fill_color = level_colors.get(level, 'FFFFFF')
                row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                
                for cell in row:
                    cell.border = thin_border
                    cell.fill = row_fill
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    if cell.column in [14, 15, 16, 17]:
                        cell.number_format = '₹#,##0.00'
            
            # Summary sheet
            summary_data = []
            for layer in range(1, 8):
                layer_data = df[df['Layer'] == layer]
                if len(layer_data) > 0:
                    summary_data.append({
                        'Layer': f'Layer {layer}',
                        'Transaction Count': len(layer_data),
                        'Total Disputed Amount': layer_data['Disputed Amount'].sum(),
                        'Total Transaction Amount': layer_data['Transaction Amount'].sum()
                    })
            
            summary_data.append({})
            summary_data.append({
                'Layer': 'TOTAL',
                'Transaction Count': len(df),
                'Total Disputed Amount': df['Disputed Amount'].sum(),
                'Total Transaction Amount': ''
            })
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            summary_ws = writer.sheets['Summary']
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            summary_ws.column_dimensions['A'].width = 20
            summary_ws.column_dimensions['B'].width = 20
            summary_ws.column_dimensions['C'].width = 25
            summary_ws.column_dimensions['D'].width = 25
        
        output.seek(0)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'Transaction_Report_{timestamp}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error generating report: {str(e)}'})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
